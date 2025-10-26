from django.views import View
from .models import Workspace
"""
Views for turnos app
"""
import logging
import json
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Avg
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import FormView, DetailView
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView,
    TemplateView, View
)
from formtools.wizard.views import SessionWizardView

from .forms import (
    EnfermeraForm, TipoTurnoForm, ConfiguracionPlanificacionForm,
    EjecucionRapidaForm,
    ImportarEnfermerasForm,
    ConfiguracionWizardStep1Form,
    ConfiguracionWizardStep2DemandaForm,
    ConfiguracionWizardStep3DurasForm,
    ConfiguracionWizardStep4BlandasForm
)
from .mixins import (
    OwnerRequiredMixin,
    FormMessageMixin, PaginationMixin, SearchMixin, FilterMixin
)
from .models import ConfiguracionPlanificacion, Ejecucion
from .models import (
    Enfermera, TipoTurno, Planilla, AsignacionTurno
)

logger = logging.getLogger(__name__)

# ========== Dashboard ==========

class DashboardView(LoginRequiredMixin, TemplateView):
    """Vista principal del dashboard"""
    template_name = 'turnos/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas generales
        stats = {
            'total_configuraciones': ConfiguracionPlanificacion.objects.count(),
            'ejecuciones_exitosas': Ejecucion.objects.filter(estado='COMPLETADA').count(),
            'enfermeras_activas': Enfermera.objects.filter(activa=True).count(),
            'dias_planificados': AsignacionTurno.objects.values('fecha').distinct().count(),
        }

        # Ejecuciones recientes (últimas 5)
        ejecuciones_recientes = Ejecucion.objects.select_related('configuracion').order_by('-fecha_inicio')[:5]

        # Actividad reciente (opcional - puedes comentar si no lo usas)
        actividad_reciente = []

        # Añadir al contexto
        context['stats'] = stats
        context['ejecuciones_recientes'] = ejecuciones_recientes
        context['actividad_reciente'] = actividad_reciente

        return context


# ========== Configuraciones ==========

class ConfiguracionListView(LoginRequiredMixin, SearchMixin, FilterMixin, PaginationMixin, ListView):
    """Lista de configuraciones"""
    model = ConfiguracionPlanificacion
    template_name = 'turnos/configuration_list.html'
    context_object_name = 'configuraciones'
    paginate_by = 12

    search_fields = ['nombre', 'descripcion']
    filter_fields = {
        'estado': 'activa',
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.select_related('creado_por').prefetch_related('enfermeras', 'turnos')

        # Filtro de estado
        estado = self.request.GET.get('estado')
        if estado == 'activa':
            queryset = queryset.filter(activa=True)
        elif estado == 'inactiva':
            queryset = queryset.filter(activa=False)

        # Ordenamiento
        orden = self.request.GET.get('orden', '-fecha_creacion')
        queryset = queryset.order_by(orden)

        return queryset


class ConfiguracionDetailView(LoginRequiredMixin, DetailView):
    """Detalle de una configuración"""
    model = ConfiguracionPlanificacion
    template_name = 'turnos/configuration_detail.html'
    context_object_name = 'configuracion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Ejecuciones recientes de esta configuración
        context['ejecuciones_recientes'] = self.object.ejecuciones.order_by('-fecha_inicio')[:5]

        return context


class ConfiguracionCreateView(LoginRequiredMixin, FormMessageMixin, CreateView):
    """Crear nueva configuración"""
    model = ConfiguracionPlanificacion
    form_class = ConfiguracionPlanificacionForm
    template_name = 'turnos/configuration_form.html'
    success_message = 'Configuración creada con éxito.'

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        return super().form_valid(form)


class ConfiguracionUpdateView(LoginRequiredMixin, OwnerRequiredMixin, FormMessageMixin, UpdateView):
    """Editar configuración existente"""
    model = ConfiguracionPlanificacion
    form_class = ConfiguracionPlanificacionForm
    template_name = 'turnos/configuration_form.html'
    success_message = 'Configuración actualizada con éxito.'
    owner_field = 'creado_por'


class ConfiguracionDeleteView(LoginRequiredMixin, OwnerRequiredMixin, DeleteView):
    """Eliminar configuración"""
    model = ConfiguracionPlanificacion
    template_name = 'turnos/configuration_confirm_delete.html'
    success_url = reverse_lazy('turnos:config_lista')
    owner_field = 'creado_por'

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Configuración eliminada con éxito.')
        return super().delete(request, *args, **kwargs)


class ConfiguracionWizardView(LoginRequiredMixin, TemplateView):
    """Wizard para crear configuración paso a paso"""
    template_name = 'turnos/config/wizard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['enfermeras'] = Enfermera.objects.filter(activa=True)
        context['turnos'] = TipoTurno.objects.filter(activo=True)
        return context


class ConfiguracionDuplicarView(LoginRequiredMixin, View):
    """Duplicar una configuración existente"""

    def post(self, request, pk):
        config_original = get_object_or_404(ConfiguracionPlanificacion, pk=pk)

        # Crear copia
        config_nueva = ConfiguracionPlanificacion.objects.create(
            nombre=f"{config_original.nombre} (Copia)",
            descripcion=config_original.descripcion,
            activa=config_original.activa,
            num_dias=config_original.num_dias,
            fecha_inicio=timezone.now().date(),
            demanda_por_turno=config_original.demanda_por_turno,
            restricciones_duras=config_original.restricciones_duras,
            restricciones_blandas=config_original.restricciones_blandas,
            num_trabajadores=config_original.num_trabajadores,
            tiempo_maximo_segundos=config_original.tiempo_maximo_segundos,
            seed=config_original.seed,
            creado_por=request.user
        )

        # Copiar relaciones ManyToMany
        config_nueva.enfermeras.set(config_original.enfermeras.all())
        config_nueva.turnos.set(config_original.turnos.all())

        messages.success(request, 'Configuración duplicada con éxito.')
        return redirect('turnos:config_detalle', pk=config_nueva.pk)


# ========== Asistente de Configuración (Wizard) ==========

FORMS = [
    ("paso1", ConfiguracionWizardStep1Form),
    ("paso2", ConfiguracionWizardStep2DemandaForm),
    ("paso3", ConfiguracionWizardStep3DurasForm),
    ("paso4", ConfiguracionWizardStep4BlandasForm),
]

TEMPLATES = {
    "paso1": "turnos/wizard/paso1_basico.html",
    "paso2": "turnos/wizard/paso2_demanda.html",
    "paso3": "turnos/wizard/paso3_duras.html",
    "paso4": "turnos/wizard/paso4_blandas.html",
}

class ConfiguracionWizardViewStepByStep(LoginRequiredMixin, SessionWizardView):
    """Wizard para crear configuración paso a paso"""
    url_name = 'turnos:config_wizard_step'

    def get_template_names(self):
        return [TEMPLATES[self.steps.current]]

    def done(self, form_list, **kwargs):
        form_data = self.get_all_cleaned_data()

        # Procesar datos JSON
        demanda_por_turno = form_data.get('demanda_por_turno')
        restricciones_duras = form_data.get('restricciones_duras')
        restricciones_blandas = form_data.get('restricciones_blandas')

        # Crear instancia del modelo
        config = ConfiguracionPlanificacion.objects.create(
            nombre=form_data['nombre'],
            descripcion=form_data.get('descripcion', ''),
            num_dias=form_data['num_dias'],
            fecha_inicio=form_data['fecha_inicio'],
            demanda_por_turno=json.loads(demanda_por_turno) if demanda_por_turno else {},
            restricciones_duras=json.loads(restricciones_duras) if restricciones_duras else [],
            restricciones_blandas=json.loads(restricciones_blandas) if restricciones_blandas else [],
            num_trabajadores=form_data.get('num_trabajadores', 4),
            tiempo_maximo_segundos=form_data.get('tiempo_maximo_segundos', 60),
            seed=form_data.get('seed'),
            creado_por=self.request.user,
            activa=True  # Activar por defecto
        )

        # Añadir relaciones ManyToMany
        config.enfermeras.set(form_data['enfermeras'])
        config.turnos.set(form_data['turnos'])

        messages.success(self.request, f'Configuración "{config.nombre}" creada con éxito.')
        return redirect('turnos:config_detalle', pk=config.pk)


# ========== Ejecuciones ==========

class EjecucionListView(LoginRequiredMixin, SearchMixin, FilterMixin, PaginationMixin, ListView):
    """Lista de ejecuciones"""
    model = Ejecucion
    template_name = 'turnos/ejecucion_list.html'
    context_object_name = 'ejecuciones'
    paginate_by = 20

    search_fields = ['configuracion__nombre']

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.select_related('configuracion', 'planilla')

        # Filtros
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)

        config_id = self.request.GET.get('proyecto_turnos')
        if config_id:
            queryset = queryset.filter(configuracion_id=config_id)

        return queryset.order_by('-fecha_inicio')


class EjecucionDetailView(LoginRequiredMixin, DetailView):
    """Vista de detalle de una ejecución"""
    model = Ejecucion
    template_name = 'turnos/ejecucion_detail.html'
    context_object_name = 'ejecucion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ejecucion = self.object

        # Si tiene planilla, calcular datos para visualización
        if ejecucion.planilla:
            planilla = ejecucion.planilla

            # Obtener todas las asignaciones
            asignaciones = planilla.asignaciones.select_related('enfermera', 'turno').order_by('fecha',
                                                                                               'enfermera__nombre')

            # Agrupar por enfermera para la tabla
            enfermeras_turnos = {}
            for asignacion in asignaciones:
                enfermera_id = asignacion.enfermera.id
                if enfermera_id not in enfermeras_turnos:
                    enfermeras_turnos[enfermera_id] = {
                        'enfermera': asignacion.enfermera,
                        'turnos': []
                    }

                enfermeras_turnos[enfermera_id]['turnos'].append({
                    'fecha': asignacion.fecha,
                    'turno': asignacion.turno,
                    'es_libre': asignacion.es_dia_libre,
                    'turno_color': self._get_turno_color(asignacion)
                })

            # Lista de días únicos
            fechas_unicas = sorted(set(a.fecha for a in asignaciones))
            dias = [{'fecha': f, 'dia_semana': f.strftime('%a')} for f in fechas_unicas]

            # Distribución de turnos por tipo
            from django.db.models import Count
            distribucion_turnos = {}

            for turno in ejecucion.configuracion.turnos.all():
                count = asignaciones.filter(turno=turno, es_dia_libre=False).count()
                distribucion_turnos[turno.nombre] = {
                    'turno': turno,
                    'cantidad': count,
                    'porcentaje': round((count / asignaciones.count() * 100), 1) if asignaciones.count() > 0 else 0
                }

            # Días libres
            dias_libres_count = asignaciones.filter(es_dia_libre=True).count()
            distribucion_turnos['LIBRE'] = {
                'turno': None,
                'cantidad': dias_libres_count,
                'porcentaje': round((dias_libres_count / asignaciones.count() * 100),
                                    1) if asignaciones.count() > 0 else 0
            }

            # Carga de trabajo por enfermera (para la tabla)
            carga_enfermeras = []
            for enfermera in ejecucion.configuracion.enfermeras.all():
                turnos_trabajados = asignaciones.filter(enfermera=enfermera, es_dia_libre=False).count()
                dias_libres = asignaciones.filter(enfermera=enfermera, es_dia_libre=True).count()
                total_dias = ejecucion.configuracion.num_dias
                porcentaje = round((turnos_trabajados / total_dias * 100), 1) if total_dias > 0 else 0

                carga_enfermeras.append({
                    'enfermera': enfermera,
                    'turnos_trabajados': turnos_trabajados,
                    'dias_libres': dias_libres,
                    'total_dias': total_dias,
                    'porcentaje': porcentaje
                })

            context['enfermeras_turnos'] = list(enfermeras_turnos.values())
            context['dias'] = dias
            context['distribucion_turnos'] = distribucion_turnos
            context['carga_enfermeras'] = carga_enfermeras

        return context

    def _get_turno_color(self, asignacion):
        """Devuelve el color del turno para el badge"""
        if asignacion.es_dia_libre:
            return 'secondary'

        if asignacion.turno:
            if asignacion.turno.nombre == 'MANANA':
                return 'warning'
            elif asignacion.turno.nombre == 'TARDE':
                return 'info'
            elif asignacion.turno.nombre == 'NOCHE':
                return 'dark'

        return 'primary'


class EjecucionDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar ejecución"""
    model = Ejecucion
    template_name = 'turnos/ejecucion_confirm_delete.html'
    success_url = reverse_lazy('turnos:ejecucion_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Ejecución eliminada con éxito.')
        return super().delete(request, *args, **kwargs)


from .tasks import ejecutar_planificacion_async


class EjecutarPlanificacionView(LoginRequiredMixin, DetailView):
    """Vista para ejecutar una planificación"""
    model = ConfiguracionPlanificacion
    template_name = 'turnos/ejecutar_planificacion.html'
    context_object_name = 'configuracion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Validaciones previas
        config = self.get_object()
        errores = []
        advertencias = []

        # Verificar enfermeras activas
        enfermeras_activas = config.enfermeras.filter(activa=True).count()
        if enfermeras_activas < 2:
            errores.append(f'Se necesitan al menos 2 enfermeras activas. Actualmente hay {enfermeras_activas}.')

        # Verificar turnos activos
        turnos_activos = config.turnos.filter(activo=True).count()
        if turnos_activos < 1:
            errores.append(f'Se necesita al menos 1 tipo de turno activo. Actualmente hay {turnos_activos}.')

        # Verificar demanda
        if not config.demanda_por_turno:
            advertencias.append(
                'No hay demanda configurada. Se usarán valores por defecto (mínimo: 1, óptimo: 2, máximo: 5).')

        # Verificar restricciones
        if not config.restricciones_duras:
            advertencias.append(
                'No hay restricciones duras configuradas. Se aplicarán restricciones básicas por defecto.')

        if not config.restricciones_blandas:
            advertencias.append(
                'No hay restricciones blandas. La planificación puede no ser óptima en términos de equidad.')

        # Verificar capacidad
        total_turnos_necesarios = config.num_dias * len(config.turnos.filter(activo=True))
        capacidad_enfermeras = enfermeras_activas * config.num_dias

        if capacidad_enfermeras < total_turnos_necesarios:
            advertencias.append(
                f'Puede no haber suficientes enfermeras. '
                f'Capacidad: {capacidad_enfermeras} turnos. '
                f'Necesarios: ~{total_turnos_necesarios} turnos.'
            )

        context['errores'] = errores
        context['advertencias'] = advertencias
        context['puede_ejecutar'] = len(errores) == 0
        context['enfermeras_activas'] = enfermeras_activas
        context['turnos_activos'] = turnos_activos

        return context

    def post(self, request, *args, **kwargs):
        """Ejecuta la planificación en segundo plano"""
        configuracion = self.get_object()

        try:
            # Crear ejecución en estado PENDIENTE
            ejecucion = Ejecucion.objects.create(
                configuracion=configuracion,
                estado='PENDIENTE'
            )

            # Ejecutar en un thread separado (para desarrollo)
            # En producción deberías usar Celery, RQ o similar
            # En lugar de threading
            ejecutar_planificacion_async.delay(ejecucion.id)

            messages.success(
                request,
                f'Ejecución #{ejecucion.pk} iniciada. La planificación se está generando en segundo plano. '
                f'Esto puede tardar hasta {configuracion.tiempo_maximo_segundos} segundos.'
            )

            return redirect('turnos:ejecucion_detalle', pk=ejecucion.pk)

        except Exception as e:
            logger.error(f"Error al crear ejecución: {str(e)}", exc_info=True)
            messages.error(
                request,
                f'Error al iniciar la ejecución: {str(e)}'
            )
            return redirect('turnos:config_detalle', pk=configuracion.pk)


class EjecucionRapidaView(LoginRequiredMixin, FormView):
    """Vista para ejecución rápida"""
    template_name = 'turnos/ejecutar_rapido.html'
    form_class = EjecucionRapidaForm
    success_url = reverse_lazy('turnos:ejecucion_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['enfermeras'] = Enfermera.objects.filter(activa=True)
        return context

    def form_valid(self, form):
        # Crear configuración rápida
        # ... implementación
        messages.success(self.request, 'Ejecución rápida iniciada.')
        return super().form_valid(form)


# ========== Enfermeras ==========

class EnfermeraListView(LoginRequiredMixin, SearchMixin, PaginationMixin, ListView):
    """Lista de enfermeras"""
    model = Enfermera
    template_name = 'turnos/enfermera_list.html'
    context_object_name = 'enfermeras'
    paginate_by = 15

    search_fields = ['nombre', 'email', 'dni']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtro de estado
        estado = self.request.GET.get('estado')
        if estado == 'activa':
            queryset = queryset.filter(activa=True)
        elif estado == 'inactiva':
            queryset = queryset.filter(activa=False)

        # Ordenamiento
        orden = self.request.GET.get('orden', 'nombre')
        queryset = queryset.order_by(orden)

        return queryset


class EnfermeraDetailView(LoginRequiredMixin, DetailView):
    """Detalle de una enfermera"""
    model = Enfermera
    template_name = 'turnos/enfermera_detail.html'
    context_object_name = 'enfermera'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas
        context['stats'] = {
            'total_turnos': AsignacionTurno.objects.filter(enfermera=self.object).count(),
            'configuraciones': ConfiguracionPlanificacion.objects.filter(enfermeras=self.object).count(),
            'turnos_noche': AsignacionTurno.objects.filter(
                enfermera=self.object,
                turno__nombre='NOCHE'
            ).count()
        }

        # Asignaciones recientes
        context['asignaciones_recientes'] = AsignacionTurno.objects.filter(
            enfermera=self.object
        ).select_related('turno', 'planilla').order_by('-fecha')[:10]

        return context


class EnfermeraCreateView(LoginRequiredMixin, FormMessageMixin, CreateView):
    """Crear nueva enfermera"""
    model = Enfermera
    form_class = EnfermeraForm
    template_name = 'turnos/enfermera_form.html'
    success_message = 'Enfermera creada con éxito.'


class EnfermeraUpdateView(LoginRequiredMixin, FormMessageMixin, UpdateView):
    """Editar enfermera"""
    model = Enfermera
    form_class = EnfermeraForm
    template_name = 'turnos/enfermera_form.html'
    success_message = 'Enfermera actualizada con éxito.'


class EnfermeraDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar enfermera"""
    model = Enfermera
    template_name = 'turnos/enfermera_confirm_delete.html'
    success_url = reverse_lazy('turnos:enfermera_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Enfermera eliminada con éxito.')
        return super().delete(request, *args, **kwargs)


class ImportarEnfermerasView(LoginRequiredMixin, FormView):
    """Importar enfermeras desde Excel"""
    template_name = 'turnos/enfermera_import.html'
    form_class = ImportarEnfermerasForm
    success_url = reverse_lazy('turnos:enfermera_lista')

    def form_valid(self, form):
        archivo = form.cleaned_data['archivo']
        sobrescribir = form.cleaned_data['sobrescribir']

        try:
            import openpyxl
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active

            creadas = 0
            actualizadas = 0
            errores = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                nombre, email, telefono, dni, activa = row[:5]

                if not nombre or not email:
                    continue

                activa = activa in ['Sí', 'Si', 'SI', 'sí', 'si', True, 1]

                # Buscar si existe
                enfermera_existente = Enfermera.objects.filter(email=email).first()

                if enfermera_existente:
                    if sobrescribir:
                        enfermera_existente.nombre = nombre
                        enfermera_existente.telefono = telefono or ''
                        enfermera_existente.dni = dni or ''
                        enfermera_existente.activa = activa
                        enfermera_existente.save()
                        actualizadas += 1
                else:
                    Enfermera.objects.create(
                        nombre=nombre,
                        email=email,
                        telefono=telefono or '',
                        dni=dni or '',
                        activa=activa
                    )
                    creadas += 1

            messages.success(
                self.request,
                f'Importación completada: {creadas} enfermeras creadas, {actualizadas} actualizadas.'
            )

        except Exception as e:
            messages.error(self.request, f'Error al importar: {str(e)}')
            return self.form_invalid(form)

        return super().form_valid(form)


class DescargarPlantillaEnfermerasView(LoginRequiredMixin, View):
    """Descarga plantilla Excel para importar enfermeras"""

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Enfermeras"

        # Headers con estilo
        headers = ['Nombre', 'Email', 'Teléfono', 'DNI', 'Activa']
        ws.append(headers)

        # Aplicar estilo a headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Ejemplo
        ws.append(['María García', 'maria@hospital.com', '600123456', '12345678A', 'Sí'])
        ws.append(['Juan López', 'juan@hospital.com', '600654321', '87654321B', 'Sí'])

        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_enfermeras.xlsx'
        wb.save(response)

        return response


# ========== Tipos de Turno ==========

class TipoTurnoListView(LoginRequiredMixin, ListView):
    """Lista de tipos de turno"""
    model = TipoTurno
    template_name = 'turnos/tipo_turno_list.html'
    context_object_name = 'tipos_turno'

    def get_queryset(self):
        return TipoTurno.objects.all().order_by('nombre')


class TipoTurnoCreateView(LoginRequiredMixin, FormMessageMixin, CreateView):
    """Crear tipo de turno"""
    model = TipoTurno
    form_class = TipoTurnoForm
    template_name = 'turnos/tipo_turno_form.html'
    success_url = reverse_lazy('turnos:tipo_turno_lista')
    success_message = 'Tipo de turno creado con éxito.'


class TipoTurnoUpdateView(LoginRequiredMixin, FormMessageMixin, UpdateView):
    """Editar tipo de turno"""
    model = TipoTurno
    form_class = TipoTurnoForm
    template_name = 'turnos/tipo_turno_form.html'
    success_url = reverse_lazy('turnos:tipo_turno_lista')
    success_message = 'Tipo de turno actualizado con éxito.'


class TipoTurnoDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar tipo de turno"""
    model = TipoTurno
    template_name = 'turnos/tipo_turno_confirm_delete.html'
    success_url = reverse_lazy('turnos:tipo_turno_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Tipo de turno eliminado con éxito.')
        return super().delete(request, *args, **kwargs)


class CrearTurnosPredeterminadosView(LoginRequiredMixin, View):
    """Crea los turnos predeterminados (Mañana, Tarde, Noche)"""

    def post(self, request):
        from datetime import time

        turnos_default = [
            {
                'nombre': 'MANANA',
                'hora_inicio': time(7, 0),
                'hora_fin': time(15, 0),
                'descripcion': 'Turno de mañana'
            },
            {
                'nombre': 'TARDE',
                'hora_inicio': time(15, 0),
                'hora_fin': time(23, 0),
                'descripcion': 'Turno de tarde'
            },
            {
                'nombre': 'NOCHE',
                'hora_inicio': time(23, 0),
                'hora_fin': time(7, 0),
                'descripcion': 'Turno de noche'
            }
        ]

        creados = 0
        for turno_data in turnos_default:
            if not TipoTurno.objects.filter(nombre=turno_data['nombre']).exists():
                TipoTurno.objects.create(**turno_data)
                creados += 1

        if creados > 0:
            messages.success(request, f'{creados} tipos de turno creados con éxito.')
        else:
            messages.info(request, 'Los turnos predeterminados ya existen.')

        return redirect('turnos:tipo_turno_lista')


# ========== Planillas ==========

class PlanillaListView(LoginRequiredMixin, SearchMixin, PaginationMixin, ListView):
    """Lista de planillas"""
    model = Planilla
    template_name = 'turnos/planilla_list.html'
    context_object_name = 'planillas'
    paginate_by = 12

    search_fields = ['nombre', 'descripcion']

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.select_related('ejecucion__configuracion').order_by('-fecha_inicio')


class PlanillaDetailView(LoginRequiredMixin, DetailView):
    """Detalle de una planilla"""
    model = Planilla
    template_name = 'turnos/planilla_detail.html'
    context_object_name = 'planilla'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener asignaciones agrupadas
        asignaciones = self.object.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera')

        context['asignaciones'] = asignaciones

        return context


class PlanillaDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar planilla"""
    model = Planilla
    template_name = 'turnos/planilla_confirm_delete.html'
    success_url = reverse_lazy('turnos:planilla_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Planilla eliminada con éxito.')
        return super().delete(request, *args, **kwargs)


# ========== Reportes ==========

class ReportesView(LoginRequiredMixin, TemplateView):
    """Vista principal de reportes"""
    template_name = 'turnos/reportes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas para los reportes
        context['stats'] = {
            'ultima_actualizacion_carga': timezone.now(),
            'ultima_actualizacion_conflictos': timezone.now(),
            'ultima_actualizacion_tendencias': timezone.now(),
            'conflictos_activos': 0,
            'total_configuraciones': ConfiguracionPlanificacion.objects.count(),
            'total_ejecuciones': Ejecucion.objects.count(),
            'total_enfermeras': Enfermera.objects.count(),
            'total_turnos_asignados': AsignacionTurno.objects.count()
        }

        return context


class ReporteCargaView(LoginRequiredMixin, TemplateView):
    """Reporte de carga de trabajo"""
    template_name = 'turnos/reporte_carga.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Filtros
        fecha_desde = self.request.GET.get('fecha_desde')
        fecha_hasta = self.request.GET.get('fecha_hasta')

        # Estadísticas por enfermera
        enfermeras = Enfermera.objects.filter(activa=True)

        datos_enfermeras = []
        for enfermera in enfermeras:
            asignaciones = AsignacionTurno.objects.filter(enfermera=enfermera)

            if fecha_desde:
                asignaciones = asignaciones.filter(fecha__gte=fecha_desde)
            if fecha_hasta:
                asignaciones = asignaciones.filter(fecha__lte=fecha_hasta)

            turnos_manana = asignaciones.filter(turno__nombre='MANANA').count()
            turnos_tarde = asignaciones.filter(turno__nombre='TARDE').count()
            turnos_noche = asignaciones.filter(turno__nombre='NOCHE').count()
            total_turnos = turnos_manana + turnos_tarde + turnos_noche

            datos_enfermeras.append({
                'enfermera': enfermera,
                'turnos_manana': turnos_manana,
                'turnos_tarde': turnos_tarde,
                'turnos_noche': turnos_noche,
                'total_turnos': total_turnos,
                'horas_totales': total_turnos * 8,  # Aproximación
                'dias_libres': asignaciones.filter(es_dia_libre=True).count(),
                'porcentaje_carga': (total_turnos / 30 * 100) if total_turnos else 0
            })

        context['datos_enfermeras'] = datos_enfermeras
        context['resumen'] = {
            'total_enfermeras': len(datos_enfermeras),
            'total_turnos': sum(d['total_turnos'] for d in datos_enfermeras),
            'horas_totales': sum(d['horas_totales'] for d in datos_enfermeras),
            'promedio_por_enfermera': sum(d['total_turnos'] for d in datos_enfermeras) / len(
                datos_enfermeras) if datos_enfermeras else 0
        }

        # Datos para gráfico
        context['distribucion_turnos'] = [
            sum(d['turnos_manana'] for d in datos_enfermeras),
            sum(d['turnos_tarde'] for d in datos_enfermeras),
            sum(d['turnos_noche'] for d in datos_enfermeras)
        ]

        return context


class ReporteConflictosView(LoginRequiredMixin, TemplateView):
    """Reporte de conflictos en planificaciones"""
    template_name = 'turnos/reporte_conflictos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Aquí se implementaría la lógica para detectar conflictos
        # Por ejemplo: turnos con descanso insuficiente, sobrecarga, etc.

        context['conflictos'] = []
        context['resumen'] = {
            'total_conflictos': 0,
            'severidad_alta': 0,
            'severidad_media': 0,
            'severidad_baja': 0
        }

        return context


class ReporteTendenciasView(LoginRequiredMixin, TemplateView):
    """Reporte de tendencias temporales"""
    template_name = 'turnos/reporte_tendencias.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Datos mensuales
        hoy = timezone.now()
        datos_mensuales = []

        for i in range(6):
            mes_inicio = hoy - timedelta(days=30 * i)
            mes_fin = mes_inicio + timedelta(days=30)

            ejecuciones_mes = Ejecucion.objects.filter(
                fecha_inicio__gte=mes_inicio,
                fecha_inicio__lt=mes_fin
            )

            datos_mensuales.append({
                'nombre': mes_inicio.strftime('%B'),
                'total': ejecuciones_mes.count(),
                'exitosas': ejecuciones_mes.filter(estado='COMPLETADA').count(),
                'fallidas': ejecuciones_mes.filter(estado='ERROR').count(),
                'tiempo_promedio': ejecuciones_mes.aggregate(Avg('duracion'))['duracion__avg'] or 0,
                'tasa_exito': (ejecuciones_mes.filter(
                    estado='COMPLETADA').count() / ejecuciones_mes.count() * 100) if ejecuciones_mes.count() > 0 else 0
            })

        context['datos_mensuales'] = datos_mensuales

        # KPIs
        context['kpis'] = {
            'crecimiento_ejecuciones': 15.5,
            'tiempo_promedio': 45.3,
            'tasa_exito': 92.5,
            'soluciones_optimas': 78.2
        }

        return context


# ========== Vistas de Usuario ==========

class PerfilView(LoginRequiredMixin, TemplateView):
    """Vista de perfil de usuario"""
    template_name = 'turnos/perfil.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas del usuario
        context['stats'] = {
            'configuraciones': ConfiguracionPlanificacion.objects.filter(creado_por=self.request.user).count(),
            'ejecuciones': Ejecucion.objects.filter(configuracion__creado_por=self.request.user).count(),
            'enfermeras': Enfermera.objects.count()
        }

        return context


class PreferenciasView(LoginRequiredMixin, TemplateView):
    """Vista de preferencias del sistema"""
    template_name = 'turnos/preferencias.html'


class GuardarPreferenciasView(LoginRequiredMixin, View):
    """Guarda las preferencias del usuario"""

    def post(self, request):
        # Guardar preferencias en la sesión o en un modelo de perfil
        request.session['preferencias'] = {
            'notificaciones_email': request.POST.get('notificaciones_email') == 'on',
            'notificaciones_browser': request.POST.get('notificaciones_browser') == 'on',
            'idioma': request.POST.get('idioma', 'es'),
            'tema': request.POST.get('tema', 'light'),
            'trabajadores_default': int(request.POST.get('trabajadores_default', 4)),
            'tiempo_maximo_default': int(request.POST.get('tiempo_maximo_default', 60))
        }

        messages.success(request, 'Preferencias guardadas con éxito.')
        return redirect('turnos:preferencias')


# ========== Vistas de Resultado ==========

class ResultadoCalendarioView(LoginRequiredMixin, DetailView):
    """Vista de resultado en formato calendario"""
    model = Ejecucion
    template_name = 'turnos/resultado_calendario.html'
    context_object_name = 'ejecucion'


class ResultadoEstadisticasView(LoginRequiredMixin, DetailView):
    """Vista de estadísticas del resultado"""
    model = Ejecucion
    template_name = 'turnos/resultado_estadisticas.html'
    context_object_name = 'ejecucion'


class ResultadoTablaView(LoginRequiredMixin, DetailView):
    """Vista de resultado en formato tabla"""
    model = Ejecucion
    template_name = 'turnos/resultado_tabla.html'
    context_object_name = 'ejecucion'


class ResultadoCompararView(LoginRequiredMixin, TemplateView):
    """Vista para comparar dos resultados"""
    template_name = 'turnos/resultado_comparar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        ejecucion_a_id = self.request.GET.get('ejecucion_a')
        ejecucion_b_id = self.request.GET.get('ejecucion_b')

        if ejecucion_a_id and ejecucion_b_id:
            context['ejecucion_a'] = get_object_or_404(Ejecucion, pk=ejecucion_a_id)
            context['ejecucion_b'] = get_object_or_404(Ejecucion, pk=ejecucion_b_id)

        context['ejecuciones'] = Ejecucion.objects.filter(estado='COMPLETADA').order_by('-fecha_inicio')[:20]

        return context


# ========== Vistas de Utilidad ==========

class MaintenanceView(TemplateView):
    """Vista de mantenimiento"""
    template_name = 'turnos/maintenance.html'


# ========== Exportaciones ==========

class ExportarEjecucionExcelView(LoginRequiredMixin, View):
    """Exporta una ejecución a Excel"""

    def get(self, request, pk):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        wb = Workbook()
        ws = wb.active
        ws.title = "Planilla de Turnos"

        # Título
        ws['A1'] = ejecucion.configuracion.nombre
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')

        ws[
            'A2'] = f"Período: {ejecucion.planilla.fecha_inicio.strftime('%d/%m/%Y')} - {ejecucion.planilla.fecha_fin.strftime('%d/%m/%Y')}"
        ws.merge_cells('A2:E2')

        # Headers
        headers = ['Enfermera', 'Fecha', 'Día Semana', 'Turno', 'Horario']
        ws.append([''])  # Línea en blanco
        ws.append(headers)

        # Estilo headers
        for cell in ws[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Datos
        asignaciones = ejecucion.planilla.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera')

        for asignacion in asignaciones:
            if asignacion.es_dia_libre:
                turno_info = 'Libre'
                horario = '-'
            else:
                turno_info = asignacion.turno.get_nombre_display()
                horario = f"{asignacion.turno.hora_inicio.strftime('%H:%M')} - {asignacion.turno.hora_fin.strftime('%H:%M')}"

            dia_semana = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom'][asignacion.fecha.weekday()]

            ws.append([
                asignacion.enfermera.nombre,
                asignacion.fecha.strftime('%d/%m/%Y'),
                dia_semana,
                turno_info,
                horario
            ])

        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20

        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.xlsx'
        wb.save(response)

        return response


class ExportarEjecucionPDFView(LoginRequiredMixin, View):
    """Exporta una ejecución a PDF"""

    def get(self, request, pk):
        from django.template.loader import render_to_string
        from weasyprint import HTML

        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        # Renderizar HTML
        html_string = render_to_string('turnos/pdf/planilla.html', {
            'ejecucion': ejecucion,
            'planilla': ejecucion.planilla,
            'asignaciones': ejecucion.planilla.asignaciones.select_related('enfermera', 'turno').order_by('fecha',
                                                                                                          'enfermera')
        })

        # Generar PDF
        html = HTML(string=html_string)
        pdf = html.write_pdf()

        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.pdf'

        return response


class ExportarEjecucionCSVView(LoginRequiredMixin, View):
    """Exporta una ejecución a CSV"""

    def get(self, request, pk):
        import csv

        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.csv'

        writer = csv.writer(response)
        writer.writerow(['Enfermera', 'Fecha', 'Turno', 'Horario', 'Es Día Libre'])

        asignaciones = ejecucion.planilla.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera')

        for asignacion in asignaciones:
            if asignacion.es_dia_libre:
                turno_info = 'Libre'
                horario = '-'
            else:
                turno_info = asignacion.turno.get_nombre_display()
                horario = f"{asignacion.turno.hora_inicio.strftime('%H:%M')} - {asignacion.turno.hora_fin.strftime('%H:%M')}"

            writer.writerow([
                asignacion.enfermera.nombre,
                asignacion.fecha.strftime('%d/%m/%Y'),
                turno_info,
                horario,
                'Sí' if asignacion.es_dia_libre else 'No'
            ])

        return response


class ExportarEjecucionJSONView(LoginRequiredMixin, View):
    """Exporta una ejecución a JSON"""

    def get(self, request, pk):
        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        data = {
            'ejecucion_id': ejecucion.id,
            'configuracion': ejecucion.configuracion.nombre,
            'fecha_inicio': ejecucion.fecha_inicio.isoformat(),
            'estado': ejecucion.estado,
            'es_optima': ejecucion.es_optima,
            'penalizacion_total': ejecucion.penalizacion_total,
            'planilla': {
                'nombre': ejecucion.planilla.nombre,
                'fecha_inicio': ejecucion.planilla.fecha_inicio.isoformat(),
                'fecha_fin': ejecucion.planilla.fecha_fin.isoformat(),
                'asignaciones': []
            }
        }

        asignaciones = ejecucion.planilla.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera')

        for asignacion in asignaciones:
            data['planilla']['asignaciones'].append({
                'enfermera': asignacion.enfermera.nombre,
                'fecha': asignacion.fecha.isoformat(),
                'turno': asignacion.turno.get_nombre_display() if asignacion.turno else None,
                'es_dia_libre': asignacion.es_dia_libre
            })

        response = JsonResponse(data, json_dumps_params={'indent': 2})
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.json'

        return response


class ExportarEjecucionICalView(LoginRequiredMixin, View):
    """Exporta una ejecución a formato iCalendar"""

    def get(self, request, pk):
        from icalendar import Calendar, Event
        from datetime import datetime, timedelta

        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        cal = Calendar()
        cal.add('prodid', '-//Sistema de Planificación de Turnos//ES')
        cal.add('version', '2.0')
        cal.add('X-WR-CALNAME', ejecucion.configuracion.nombre)

        asignaciones = ejecucion.planilla.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha')

        for asignacion in asignaciones:
            if not asignacion.es_dia_libre:
                event = Event()
                event.add('summary', f"{asignacion.enfermera.nombre} - {asignacion.turno.get_nombre_display()}")

                # Calcular fecha/hora inicio y fin
                dt_inicio = datetime.combine(asignacion.fecha, asignacion.turno.hora_inicio)
                dt_fin = datetime.combine(asignacion.fecha, asignacion.turno.hora_fin)

                # Si el turno cruza medianoche
                if dt_fin <= dt_inicio:
                    dt_fin += timedelta(days=1)

                event.add('dtstart', dt_inicio)
                event.add('dtend', dt_fin)
                event.add('description', f"Turno asignado: {asignacion.turno.get_nombre_display()}")

                cal.add_component(event)

        response = HttpResponse(cal.to_ical(), content_type='text/calendar')
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.ics'

        return response


class ExportarPlanillaExcelView(LoginRequiredMixin, View):
    """Exporta una planilla específica a Excel"""

    def get(self, request, pk):
        planilla = get_object_or_404(Planilla, pk=pk)
        # Reutilizar la lógica de exportación de ejecución
        # pero usando la planilla directamente
        return ExportarEjecucionExcelView().get(request, planilla.ejecucion.id)


class ExportarPlanillaPDFView(LoginRequiredMixin, View):
    """Exporta una planilla específica a PDF"""

    def get(self, request, pk):
        planilla = get_object_or_404(Planilla, pk=pk)
        return ExportarEjecucionPDFView().get(request, planilla.ejecucion.id)

class ExportarConfiguracionJSONView(LoginRequiredMixin, DetailView):
    """Vista para exportar configuración en formato JSON"""
    model = ConfiguracionPlanificacion

    def get(self, request, *args, **kwargs):
        configuracion = self.get_object()

        data = {
            'nombre': configuracion.nombre,
            'descripcion': configuracion.descripcion,
            'num_dias': configuracion.num_dias,
            'fecha_inicio': configuracion.fecha_inicio.isoformat() if configuracion.fecha_inicio else None,
            'enfermeras': list(configuracion.enfermeras.values_list('id', flat=True)),
            'turnos': list(configuracion.turnos.values_list('id', flat=True)),
            'demanda_por_turno': configuracion.demanda_por_turno,
            'restricciones_duras': configuracion.restricciones_duras,
            'restricciones_blandas': configuracion.restricciones_blandas,
            'num_trabajadores': configuracion.num_trabajadores,
            'tiempo_maximo_segundos': configuracion.tiempo_maximo_segundos,
            'seed': configuracion.seed,
        }

        response = JsonResponse(data, json_dumps_params={'indent': 2, 'ensure_ascii': False})
        response['Content-Disposition'] = f'attachment; filename="config_{configuracion.pk}.json"'
        return response

class WorkspaceMixin(LoginRequiredMixin):
    def get_current_workspace(self):
        workspace_id = self.request.session.get('workspace_id')
        if workspace_id:
            return get_object_or_404(Workspace, id=workspace_id, usuarios=self.request.user)
        return self.request.user.workspaces.first()

    def get_queryset(self):
        qs = super().get_queryset()
        ws = self.get_current_workspace()
        if ws:
            return qs.filter(workspace=ws)
        return qs.none()


class CambiarWorkspaceView(LoginRequiredMixin, View):
    def post(self, request):
        workspace_id = request.POST.get('workspace_id')
        ws = get_object_or_404(Workspace, id=workspace_id, usuarios=request.user)
        request.session['workspace_id'] = ws.id
        return JsonResponse({'success': True, 'workspace': ws.nombre})

# ══════════════════════════════════════════════════════════
# VISTAS PARA GESTIONAR RESTRICCIONES
# ══════════════════════════════════════════════════════════

from django.views.generic import TemplateView
from django.contrib import messages
from django.shortcuts import redirect
from .forms import (
    ConfiguracionPlanificacionForm,
    RestriccionDuraForm,
    RestriccionBlandaForm,
    CargarRestriccionesSACYLForm
)
import json


class ConfiguracionRestriccionesView(LoginRequiredMixin, TemplateView):
    """Vista para editar restricciones de una configuración"""
    template_name = 'turnos/restricciones_form.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        config_id = self.kwargs.get('pk')

        try:
            config = ConfiguracionPlanificacion.objects.get(pk=config_id)
            ctx['config'] = config
            ctx['restricciones_duras'] = config.restricciones_duras or []
            ctx['restricciones_blandas'] = config.restricciones_blandas or []
            ctx['form_dura'] = RestriccionDuraForm()
            ctx['form_blanda'] = RestriccionBlandaForm()
            ctx['form_cargar'] = CargarRestriccionesSACYLForm()
        except ConfiguracionPlanificacion.DoesNotExist:
            messages.error(self.request, 'Configuración no encontrada')

        return ctx

    def post(self, request, *args, **kwargs):
        config_id = self.kwargs.get('pk')
        action = request.POST.get('action')

        try:
            config = ConfiguracionPlanificacion.objects.get(pk=config_id)
        except ConfiguracionPlanificacion.DoesNotExist:
            messages.error(request, 'Configuración no encontrada')
            return redirect('turnos:config_lista')

        # Añadir restricción dura
        if action == 'add_dura':
            form = RestriccionDuraForm(request.POST)
            if form.is_valid():
                rd = config.restricciones_duras or []
                if not isinstance(rd, list):
                    rd = []
                rd.append(form.to_dict())
                config.restricciones_duras = rd
                config.save()
                messages.success(request, f'Restricción dura "{form.cleaned_data["nombre"]}" añadida')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')

        # Añadir restricción blanda
        elif action == 'add_blanda':
            form = RestriccionBlandaForm(request.POST)
            if form.is_valid():
                rb = config.restricciones_blandas or []
                if not isinstance(rb, list):
                    rb = []
                rb.append(form.to_dict())
                config.restricciones_blandas = rb
                config.save()
                messages.success(request, f'Restricción blanda "{form.cleaned_data["nombre"]}" añadida')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')

        # Cargar preset SACYL
        elif action == 'cargar_sacyl':
            form = CargarRestriccionesSACYLForm(request.POST, request.FILES)
            if form.is_valid():
                preset = form.cleaned_data['preset']

                if preset == 'basico':
                    # Cargar preset básico hardcoded
                    config.restricciones_duras = [
                        {"id": "RD006", "nombre": "descanso_minimo_12h", "tipo": "descanso", "obligatorio": True, "parametros": {"minimo_horas": 12}, "descripcion": "Descanso mínimo 12h entre jornadas"},
                        {"id": "RD019", "nombre": "cobertura_minima", "tipo": "cobertura", "obligatorio": True, "parametros": {}, "descripcion": "Cobertura mínima por turno"},
                        {"id": "RD020", "nombre": "no_solapamiento", "tipo": "asignacion", "obligatorio": True, "parametros": {}, "descripcion": "Una enfermera, un turno por día"}
                    ]
                    config.restricciones_blandas = [
                        {"id": "RB001", "nombre": "equidad_turnos", "tipo": "equidad", "peso": 100, "parametros": {}, "descripcion": "Distribución equitativa de turnos"}
                    ]
                    config.save()
                    messages.success(request, 'Preset básico SACYL cargado')

                elif preset == 'custom' and 'json_data' in form.cleaned_data:
                    data = form.cleaned_data['json_data']
                    if 'restricciones_duras' in data:
                        config.restricciones_duras = data['restricciones_duras']
                    if 'restricciones_blandas' in data:
                        config.restricciones_blandas = data['restricciones_blandas']
                    config.save()
                    messages.success(request, 'Restricciones personalizadas cargadas desde JSON')
            else:
                for error in form.non_field_errors():
                    messages.error(request, error)

        # Eliminar restricción
        elif action == 'delete_dura':
            idx = int(request.POST.get('index', -1))
            if idx >= 0:
                rd = config.restricciones_duras or []
                if isinstance(rd, list) and idx < len(rd):
                    del rd[idx]
                    config.restricciones_duras = rd
                    config.save()
                    messages.success(request, 'Restricción dura eliminada')

        elif action == 'delete_blanda':
            idx = int(request.POST.get('index', -1))
            if idx >= 0:
                rb = config.restricciones_blandas or []
                if isinstance(rb, list) and idx < len(rb):
                    del rb[idx]
                    config.restricciones_blandas = rb
                    config.save()
                    messages.success(request, 'Restricción blanda eliminada')

        return redirect('turnos:config_restricciones', pk=config_id)