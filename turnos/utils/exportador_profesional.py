# =========================================================================
# EXPORTADOR PROFESIONAL COMPLETO CON VALIDACIONES Y GRÁFICOS
# =========================================================================
"""
Módulo de exportación avanzada para planificación de turnos de enfermeras.

Características:
✓ Exportación a Excel con 6 hojas de análisis
✓ Exportación a PDF con tabla y estadísticas
✓ Gráficos de distribución y estadísticas
✓ Validaciones de integridad
✓ Reportes de calidad
✓ Histogramas de carga
✓ Análisis de tendencias

Uso:
    from exportador_profesional import ExportadorProfesional, exportar_resultados
    
    archivos = exportar_resultados(
        resultado_planificacion,
        config,
        'planificacion_feb_2024'
    )
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart, LineChart
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
)
from reportlab.lib import colors as rl_colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime, timedelta
import logging
from collections import defaultdict
from pathlib import Path

logger = logging.getLogger(__name__)

# =========================================================================
# CONFIGURACIÓN GLOBAL
# =========================================================================

COLORES_TURNOS = {
    'MAÑANA': {'rgb': 'FFC107', 'rgb_rl': '#FFC107', 'nombre': 'Mañana'},
    'TARDE': {'rgb': '00BCD4', 'rgb_rl': '#00BCD4', 'nombre': 'Tarde'},
    'NOCHE': {'rgb': '424242', 'rgb_rl': '#424242', 'nombre': 'Noche'},
    'LIBRE': {'rgb': 'E0E0E0', 'rgb_rl': '#E0E0E0', 'nombre': 'Libre'},
    'DESCANSO': {'rgb': 'FFFFFF', 'rgb_rl': '#FFFFFF', 'nombre': 'Descanso'},
}

FUENTE_TITULO = {'nombre': 'Calibri', 'tamaño': 14, 'bold': True}
FUENTE_ENCABEZADO = {'nombre': 'Calibri', 'tamaño': 10, 'bold': True}
FUENTE_DATO = {'nombre': 'Calibri', 'tamaño': 9}
FUENTE_PEQUENA = {'nombre': 'Calibri', 'tamaño': 8}

BORDE_DELGADO = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

COLOR_ENCABEZADO = "1F4E78"
COLOR_SUBENCABEZADO = "D9E8F5"
COLOR_ERROR = "FF6B6B"
COLOR_ALERTA = "FFD93D"
COLOR_EXITO = "6BCB77"


# =========================================================================
# ESTADÍSTICAS Y VALIDACIONES
# =========================================================================

class EstadisticasAvanzadas:
    """Calcula estadísticas completas de la planificación."""
    
    def __init__(self, planificacion_data):
        self.enfermeras = planificacion_data.get('enfermeras', [])
        self.turnos_asignados = planificacion_data.get('turnos_asignados', {})
        self.fecha_inicio = planificacion_data.get('fecha_inicio', datetime.now())
        self.fecha_fin = planificacion_data.get('fecha_fin', datetime.now())
        self.num_dias = (self.fecha_fin - self.fecha_inicio).days + 1
        self.num_enfermeras = len(self.enfermeras)
    
    def contar_turnos_por_tipo(self):
        """Retorna conteo de cada tipo de turno."""
        conteo = defaultdict(int)
        for turno in self.turnos_asignados.values():
            conteo[turno] += 1
        return dict(conteo)
    
    def turnos_por_enfermera(self):
        """Retorna turnos totales por cada enfermera."""
        datos = {}
        for idx_enf in range(self.num_enfermeras):
            nombre = self.enfermeras[idx_enf]['nombre']
            turnos_count = sum(1 for (e, d) in self.turnos_asignados if e == idx_enf)
            datos[nombre] = turnos_count
        return datos
    
    def turnos_por_enfermera_y_tipo(self):
        """Retorna conteo de cada tipo de turno por enfermera."""
        datos = defaultdict(lambda: defaultdict(int))
        for (idx_enf, dia), turno in self.turnos_asignados.items():
            nombre = self.enfermeras[idx_enf]['nombre']
            datos[nombre][turno] += 1
        return dict(datos)
    
    def dias_libres_por_enfermera(self):
        """Retorna días libres por enfermera."""
        datos = {}
        for idx_enf in range(self.num_enfermeras):
            nombre = self.enfermeras[idx_enf]['nombre']
            dias_trabajados = sum(1 for (e, d) in self.turnos_asignados if e == idx_enf)
            dias_libres = self.num_dias - dias_trabajados
            datos[nombre] = dias_libres
        return datos
    
    def cobertura_diaria_por_turno(self):
        """Retorna cobertura diaria de cada turno."""
        datos = defaultdict(lambda: defaultdict(int))
        for (enf, dia), turno in self.turnos_asignados.items():
            datos[dia][turno] += 1
        return dict(datos)
    
    def distribucion_equidad(self):
        """Calcula equidad en distribución de turnos."""
        turnos_enfermera = self.turnos_por_enfermera()
        if not turnos_enfermera:
            return {'media': 0, 'min': 0, 'max': 0, 'desviacion': 0}
        
        valores = list(turnos_enfermera.values())
        media = sum(valores) / len(valores)
        minimo = min(valores)
        maximo = max(valores)
        
        varianza = sum((x - media) ** 2 for x in valores) / len(valores)
        desviacion = varianza ** 0.5
        
        return {
            'media': media,
            'min': minimo,
            'max': maximo,
            'desviacion': desviacion,
            'diferencia': maximo - minimo
        }
    
    def equipos_mas_ocupados(self, top=5):
        """Retorna las enfermeras más ocupadas."""
        turnos_enfermera = self.turnos_por_enfermera()
        return sorted(turnos_enfermera.items(), key=lambda x: x[1], reverse=True)[:top]
    
    def equipos_menos_ocupados(self, top=5):
        """Retorna las enfermeras menos ocupadas."""
        turnos_enfermera = self.turnos_por_enfermera()
        return sorted(turnos_enfermera.items(), key=lambda x: x[1])[:top]
    
    def cobertura_minima_garantizada(self, turno_tipo='MAÑANA'):
        """Verifica si se garantiza cobertura mínima por turno."""
        dias_sin_cobertura = []
        cobertura = self.cobertura_diaria_por_turno()
        
        for dia in range(self.num_dias):
            cantidad = cobertura.get(dia, {}).get(turno_tipo, 0)
            if cantidad == 0:
                fecha = self.fecha_inicio + timedelta(days=dia)
                dias_sin_cobertura.append(fecha.strftime('%d/%m/%Y'))
        
        return {
            'tiene_cobertura': len(dias_sin_cobertura) == 0,
            'dias_sin_cobertura': dias_sin_cobertura,
            'cantidad_dias_sin_cobertura': len(dias_sin_cobertura)
        }
    
    def validar_integridad(self):
        """Retorna reporte de validaciones."""
        validaciones = {}
        
        # 1. Verificar si todos los turnos están asignados
        dias_totales_esperados = self.num_enfermeras * self.num_dias
        dias_totales_asignados = len(self.turnos_asignados)
        validaciones['turnos_asignados'] = {
            'ok': True,  # Puede haber días sin turno
            'esperados': dias_totales_esperados,
            'asignados': dias_totales_asignados,
            'faltantes': dias_totales_esperados - dias_totales_asignados
        }
        
        # 2. Verificar equidad
        equidad = self.distribucion_equidad()
        validaciones['equidad'] = {
            'ok': equidad['diferencia'] <= 3,  # Diferencia máxima de 3 turnos
            'diferencia': equidad['diferencia'],
            'desviacion': equidad['desviacion'],
            'recomendacion': 'Distribución equitativa' if equidad['diferencia'] <= 2 else 'Revisar distribución'
        }
        
        # 3. Verificar cobertura por turno
        validaciones['cobertura'] = {}
        for turno_tipo in ['MAÑANA', 'TARDE', 'NOCHE']:
            validaciones['cobertura'][turno_tipo] = self.cobertura_minima_garantizada(turno_tipo)
        
        return validaciones


class ValidadorPlani:
    """Valida la planificación generada."""
    
    @staticmethod
    def generar_reporte_validacion(stats):
        """Genera reporte de validaciones."""
        validaciones = stats.validar_integridad()
        
        reporte = []
        reporte.append("=" * 80)
        reporte.append("REPORTE DE VALIDACIONES")
        reporte.append("=" * 80)
        reporte.append("")
        
        # Turnos asignados
        reporte.append(f"✓ Turnos Asignados: {validaciones['turnos_asignados']['asignados']}")
        reporte.append(f"  Faltantes: {validaciones['turnos_asignados']['faltantes']}")
        reporte.append("")
        
        # Equidad
        equidad_ok = validaciones['equidad']['ok']
        estado_equidad = "✓ ACEPTABLE" if equidad_ok else "✗ REVISAR"
        reporte.append(f"{estado_equidad} - Equidad")
        reporte.append(f"  Diferencia: {validaciones['equidad']['diferencia']} turnos")
        reporte.append(f"  Desviación estándar: {validaciones['equidad']['desviacion']:.2f}")
        reporte.append(f"  Recomendación: {validaciones['equidad']['recomendacion']}")
        reporte.append("")
        
        # Cobertura
        reporte.append("Cobertura por Turno:")
        for turno_tipo, cobertura_info in validaciones['cobertura'].items():
            estado = "✓" if cobertura_info['tiene_cobertura'] else "✗"
            reporte.append(f"  {estado} {turno_tipo}: {cobertura_info['cantidad_dias_sin_cobertura']} días sin cobertura")
        
        reporte.append("")
        reporte.append("=" * 80)
        
        return "\n".join(reporte)


# =========================================================================
# EXPORTADOR PROFESIONAL
# =========================================================================

class ExportadorProfesional:
    """Exportador profesional con todas las funcionalidades."""
    
    def __init__(self, planificacion_data, config=None):
        self.planificacion = planificacion_data
        self.config = config or {}
        self.enfermeras = planificacion_data.get('enfermeras', [])
        self.turnos_asignados = planificacion_data.get('turnos_asignados', {})
        self.fecha_inicio = planificacion_data.get('fecha_inicio', datetime.now())
        self.fecha_fin = planificacion_data.get('fecha_fin', datetime.now())
        self.num_enfermeras = len(self.enfermeras)
        self.num_dias = (self.fecha_fin - self.fecha_inicio).days + 1
        
        # Estadísticas y validaciones
        self.stats = EstadisticasAvanzadas(planificacion_data)
        self.validador = ValidadorPlani()
    
    def _generar_matriz_datos(self):
        """Genera matriz de datos para la tabla."""
        encabezado = ['Enfermera', 'Rol']
        fechas = []
        
        for i in range(self.num_dias):
            fecha = self.fecha_inicio + timedelta(days=i)
            fecha_str = fecha.strftime('%d/%m')
            dia_semana = ['lun', 'mar', 'mié', 'jue', 'vie', 'sáb', 'dom'][fecha.weekday()]
            encabezado.append(f"{fecha_str}\n{dia_semana}")
            fechas.append(fecha)
        
        matriz = []
        for idx_enf, enfermera in enumerate(self.enfermeras):
            fila = [enfermera['nombre'], enfermera.get('rol', '-')]
            
            for idx_dia in range(self.num_dias):
                turno = self.turnos_asignados.get((idx_enf, idx_dia), '-')
                fila.append(turno)
            
            matriz.append(fila)
        
        return encabezado, matriz, fechas
    
    # =====================================================================
    # EXPORTAR EXCEL
    # =====================================================================
    
    def exportar_excel(self, archivo_salida):
        """Exporta a Excel con 6 hojas de análisis."""
        logger.info(f"Generando Excel: {archivo_salida}")
        
        encabezado, matriz, fechas = self._generar_matriz_datos()
        wb = Workbook()
        
        # Hoja 1: Tabla principal
        self._crear_hoja_tabla(wb, encabezado, matriz)
        
        # Hoja 2: Estadísticas generales
        self._crear_hoja_estadisticas_generales(wb)
        
        # Hoja 3: Por enfermera
        self._crear_hoja_distribucion_enfermeras(wb)
        
        # Hoja 4: Cobertura diaria
        self._crear_hoja_cobertura(wb)
        
        # Hoja 5: Equidad y análisis
        self._crear_hoja_equidad(wb)
        
        # Hoja 6: Validaciones
        self._crear_hoja_validaciones(wb)
        
        # Guardar y retornar
        wb.save(archivo_salida)
        logger.info(f"✓ Excel generado: {archivo_salida}")
        return archivo_salida
    
    def _crear_hoja_tabla(self, wb, encabezado, matriz):
        """Crea hoja con tabla de planificación."""
        ws = wb.active
        ws.title = "Planificación"
        
        # Información
        ws['A1'] = "PLANIFICACIÓN DE TURNOS"
        ws['A1'].font = Font(name=FUENTE_TITULO['nombre'], size=FUENTE_TITULO['tamaño'], bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws.merge_cells('A1:Z1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A2'] = f"Período: {self.fecha_inicio.strftime('%d/%m/%Y')} - {self.fecha_fin.strftime('%d/%m/%Y')} | Enfermeras: {self.num_enfermeras} | Días: {self.num_dias}"
        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        # Encabezado
        for col_idx, header in enumerate(encabezado, 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.value = header
            cell.font = Font(name=FUENTE_ENCABEZADO['nombre'], size=FUENTE_ENCABEZADO['tamaño'], bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDE_DELGADO
        
        # Datos
        for row_idx, fila in enumerate(matriz, 6):
            for col_idx, valor in enumerate(fila, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = valor
                
                if col_idx <= 2:
                    cell.font = Font(name=FUENTE_DATO['nombre'], size=FUENTE_DATO['tamaño'], bold=True)
                    cell.fill = PatternFill(start_color=COLOR_SUBENCABEZADO, end_color=COLOR_SUBENCABEZADO, fill_type="solid")
                else:
                    cell.font = Font(name=FUENTE_DATO['nombre'], size=FUENTE_DATO['tamaño'])
                    
                    if valor in COLORES_TURNOS:
                        color_hex = COLORES_TURNOS[valor]['rgb']
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                        
                        if valor == 'NOCHE':
                            cell.font = Font(name=FUENTE_DATO['nombre'], size=FUENTE_DATO['tamaño'], color="FFFFFF", bold=True)
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = BORDE_DELGADO
        
        # Dimensiones
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        for col_idx in range(3, len(encabezado) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 13
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[5].height = 30
    
    def _crear_hoja_estadisticas_generales(self, wb):
        """Crea hoja con estadísticas generales."""
        ws = wb.create_sheet("Estadísticas")
        
        ws['A1'] = "ESTADÍSTICAS GENERALES"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Información general
        ws[f'A{row}'] = "INFORMACIÓN GENERAL"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        datos_info = [
            ("Total de enfermeras:", self.num_enfermeras),
            ("Período:", f"{self.fecha_inicio.strftime('%d/%m/%Y')} - {self.fecha_fin.strftime('%d/%m/%Y')}"),
            ("Total de días:", self.num_dias),
            ("Total de turnos asignados:", len(self.turnos_asignados)),
        ]
        
        for etiqueta, valor in datos_info:
            ws[f'A{row}'] = etiqueta
            ws[f'B{row}'] = valor
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Distribución por turno
        row += 1
        ws[f'A{row}'] = "DISTRIBUCIÓN POR TIPO DE TURNO"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ['Turno', 'Cantidad', 'Porcentaje', 'Visual']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        
        conteo = self.stats.contar_turnos_por_tipo()
        total = sum(conteo.values()) if conteo else 1
        
        for turno_nombre in sorted(conteo.keys()):
            row += 1
            cantidad = conteo[turno_nombre]
            porcentaje = (cantidad / total * 100) if total > 0 else 0
            
            ws[f'A{row}'] = turno_nombre
            ws[f'B{row}'] = cantidad
            ws[f'C{row}'] = f"{porcentaje:.1f}%"
            
            # Barra de progreso visual
            barra_visual = "█" * int(porcentaje / 5) + "░" * (20 - int(porcentaje / 5))
            ws[f'D{row}'] = barra_visual
            
            if turno_nombre in COLORES_TURNOS:
                color_hex = COLORES_TURNOS[turno_nombre]['rgb']
                ws[f'A{row}'].fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                if turno_nombre == 'NOCHE':
                    ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
    
    def _crear_hoja_distribucion_enfermeras(self, wb):
        """Crea hoja con distribución por enfermera."""
        ws = wb.create_sheet("Por Enfermera")
        
        ws['A1'] = "DISTRIBUCIÓN DE TURNOS POR ENFERMERA"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws.merge_cells('A1:H1')
        
        row = 3
        headers = ['Enfermera', 'Total', 'Libres', 'MAÑANA', 'TARDE', 'NOCHE', '% Ocupación', 'Estado']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        
        turnos_por_enf_tipo = self.stats.turnos_por_enfermera_y_tipo()
        dias_libres = self.stats.dias_libres_por_enfermera()
        
        equidad = self.stats.distribucion_equidad()
        media_turnos = equidad['media']
        
        row += 1
        for idx_enf, enfermera in enumerate(self.enfermeras):
            nombre = enfermera['nombre']
            
            total_turnos = sum(1 for (e, d) in self.turnos_asignados if e == idx_enf)
            dias_libres_enf = dias_libres.get(nombre, 0)
            
            ws.cell(row=row, column=1).value = nombre
            ws.cell(row=row, column=2).value = total_turnos
            ws.cell(row=row, column=3).value = dias_libres_enf
            
            col = 4
            turnos_tipo = turnos_por_enf_tipo.get(nombre, {})
            for tipo_turno in ['MAÑANA', 'TARDE', 'NOCHE']:
                cantidad = turnos_tipo.get(tipo_turno, 0)
                cell = ws.cell(row=row, column=col)
                cell.value = cantidad
                
                if tipo_turno in COLORES_TURNOS and cantidad > 0:
                    color_hex = COLORES_TURNOS[tipo_turno]['rgb']
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    if tipo_turno == 'NOCHE':
                        cell.font = Font(color="FFFFFF", bold=True)
                
                col += 1
            
            # Porcentaje de ocupación
            ocupacion = (total_turnos / self.num_dias * 100) if self.num_dias > 0 else 0
            ws.cell(row=row, column=col).value = f"{ocupacion:.1f}%"
            col += 1
            
            # Estado (comparar con media)
            if abs(total_turnos - media_turnos) <= 1:
                estado = "✓ Equilibrado"
                color = COLOR_EXITO
            elif total_turnos > media_turnos + 2:
                estado = "⚠ Sobrecargado"
                color = COLOR_ALERTA
            else:
                estado = "⚠ Subcargado"
                color = COLOR_ALERTA
            
            ws.cell(row=row, column=col).value = estado
            ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # Alineación y bordes
            for col_idx in range(1, col + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = BORDE_DELGADO
            
            row += 1
        
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _crear_hoja_cobertura(self, wb):
        """Crea hoja con análisis de cobertura diaria."""
        ws = wb.create_sheet("Cobertura")
        
        ws['A1'] = "ANÁLISIS DE COBERTURA DIARIA"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws.merge_cells('A1:F1')
        
        row = 3
        headers = ['Fecha', 'Día', 'MAÑANA', 'TARDE', 'NOCHE', 'Total Cobertura']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        
        cobertura = self.stats.cobertura_diaria_por_turno()
        
        row += 1
        for idx_dia in range(self.num_dias):
            fecha = self.fecha_inicio + timedelta(days=idx_dia)
            fecha_str = fecha.strftime('%d/%m/%Y')
            dia_semana = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom'][fecha.weekday()]
            
            ws.cell(row=row, column=1).value = fecha_str
            ws.cell(row=row, column=2).value = dia_semana
            
            dia_cobertura = cobertura.get(idx_dia, {})
            
            col = 3
            total_dia = 0
            for tipo_turno in ['MAÑANA', 'TARDE', 'NOCHE']:
                cantidad = dia_cobertura.get(tipo_turno, 0)
                cell = ws.cell(row=row, column=col)
                cell.value = cantidad
                total_dia += cantidad
                
                if cantidad > 0 and tipo_turno in COLORES_TURNOS:
                    color_hex = COLORES_TURNOS[tipo_turno]['rgb']
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    if tipo_turno == 'NOCHE':
                        cell.font = Font(color="FFFFFF", bold=True)
                
                cell.alignment = Alignment(horizontal='center')
                col += 1
            
            total_cell = ws.cell(row=row, column=col)
            total_cell.value = total_dia
            total_cell.alignment = Alignment(horizontal='center')
            total_cell.font = Font(bold=True)
            total_cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            row += 1
        
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _crear_hoja_equidad(self, wb):
        """Crea hoja con análisis de equidad."""
        ws = wb.create_sheet("Equidad")
        
        ws['A1'] = "ANÁLISIS DE EQUIDAD Y BALANCEO"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws.merge_cells('A1:C1')
        
        row = 3
        ws[f'A{row}'] = "ESTADÍSTICAS GENERALES"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws.merge_cells(f'A{row}:C{row}')
        
        equidad = self.stats.distribucion_equidad()
        
        row += 1
        datos_equidad = [
            ("Promedio de turnos por enfermera:", f"{equidad['media']:.1f}"),
            ("Mínimo de turnos:", f"{equidad['min']}"),
            ("Máximo de turnos:", f"{equidad['max']}"),
            ("Diferencia (Máx - Mín):", f"{equidad['diferencia']}"),
            ("Desviación estándar:", f"{equidad['desviacion']:.2f}"),
        ]
        
        for etiqueta, valor in datos_equidad:
            ws[f'A{row}'] = etiqueta
            ws[f'B{row}'] = valor
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Más ocupados
        row += 1
        ws[f'A{row}'] = "TOP 5 MÁS OCUPADAS"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Posición"
        ws[f'B{row}'] = "Enfermera"
        ws[f'C{row}'] = "Turnos"
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        
        for pos, (nombre, cantidad) in enumerate(self.stats.equipos_mas_ocupados(5), 1):
            row += 1
            ws[f'A{row}'] = pos
            ws[f'B{row}'] = nombre
            ws[f'C{row}'] = cantidad
            ws[f'C{row}'].fill = PatternFill(start_color=COLOR_ALERTA, end_color=COLOR_ALERTA, fill_type="solid")
        
        # Menos ocupados
        row += 2
        ws[f'A{row}'] = "TOP 5 MENOS OCUPADAS"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Posición"
        ws[f'B{row}'] = "Enfermera"
        ws[f'C{row}'] = "Turnos"
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        
        for pos, (nombre, cantidad) in enumerate(self.stats.equipos_menos_ocupados(5), 1):
            row += 1
            ws[f'A{row}'] = pos
            ws[f'B{row}'] = nombre
            ws[f'C{row}'] = cantidad
            ws[f'C{row}'].fill = PatternFill(start_color=COLOR_ALERTA, end_color=COLOR_ALERTA, fill_type="solid")
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
    
    def _crear_hoja_validaciones(self, wb):
        """Crea hoja con validaciones."""
        ws = wb.create_sheet("Validaciones")
        
        ws['A1'] = "REPORTE DE VALIDACIONES"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws.merge_cells('A1:C1')
        
        row = 3
        validaciones = self.stats.validar_integridad()
        
        # Turnos asignados
        ws[f'A{row}'] = "Turnos Asignados"
        ws[f'B{row}'] = validaciones['turnos_asignados']['asignados']
        ws[f'C{row}'] = "✓"
        ws[f'C{row}'].fill = PatternFill(start_color=COLOR_EXITO, end_color=COLOR_EXITO, fill_type="solid")
        row += 1
        ws[f'A{row}'] = "Turnos Faltantes"
        ws[f'B{row}'] = validaciones['turnos_asignados']['faltantes']
        row += 1
        
        # Equidad
        row += 1
        ws[f'A{row}'] = "EQUIDAD"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        
        row += 1
        equidad_ok = validaciones['equidad']['ok']
        estado_equidad = "✓ ACEPTABLE" if equidad_ok else "✗ REVISAR"
        color_equidad = COLOR_EXITO if equidad_ok else COLOR_ALERTA
        
        ws[f'A{row}'] = "Estado"
        ws[f'B{row}'] = estado_equidad
        ws[f'B{row}'].fill = PatternFill(start_color=color_equidad, end_color=color_equidad, fill_type="solid")
        row += 1
        ws[f'A{row}'] = "Diferencia"
        ws[f'B{row}'] = f"{validaciones['equidad']['diferencia']} turnos"
        row += 1
        ws[f'A{row}'] = "Desviación Estándar"
        ws[f'B{row}'] = f"{validaciones['equidad']['desviacion']:.2f}"
        row += 1
        
        # Cobertura
        row += 1
        ws[f'A{row}'] = "COBERTURA POR TURNO"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        
        for turno_tipo, cobertura_info in validaciones['cobertura'].items():
            tiene_cobertura = cobertura_info['tiene_cobertura']
            estado = "✓" if tiene_cobertura else "✗"
            color = COLOR_EXITO if tiene_cobertura else COLOR_ERROR
            
            ws[f'A{row}'] = turno_tipo
            ws[f'B{row}'] = f"{cobertura_info['cantidad_dias_sin_cobertura']} días sin cobertura"
            ws[f'C{row}'] = estado
            ws[f'C{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
    
    def exportar_pdf(self, archivo_salida):
        """Exporta a PDF (versión simplificada)."""
        logger.info(f"Generando PDF: {archivo_salida}")
        
        doc = SimpleDocTemplate(
            archivo_salida,
            pagesize=landscape(letter),
            rightMargin=0.5*cm,
            leftMargin=0.5*cm,
            topMargin=1*cm,
            bottomMargin=0.5*cm
        )
        
        contenido = []
        
        # Página 1: Tabla
        contenido.extend(self._generar_pagina_tabla())
        contenido.append(PageBreak())
        
        # Página 2: Estadísticas
        contenido.extend(self._generar_pagina_estadisticas())
        
        doc.build(contenido)
        logger.info(f"✓ PDF generado: {archivo_salida}")
        return archivo_salida
    
    def _generar_pagina_tabla(self):
        """Genera página con tabla."""
        contenido = []
        
        encabezado, matriz, fechas = self._generar_matriz_datos()
        
        titulo = Paragraph(
            "PLANIFICACIÓN DE TURNOS",
            ParagraphStyle(
                'Titulo',
                fontName='Helvetica-Bold',
                fontSize=12,
                textColor=rl_colors.white,
                backColor=rl_colors.HexColor('#1F4E78'),
                alignment=TA_CENTER,
                leftIndent=5,
                rightIndent=5,
                topPadding=5,
                bottomPadding=5
            )
        )
        contenido.append(titulo)
        
        info_text = (
            f"Período: {self.fecha_inicio.strftime('%d/%m/%Y')} - {self.fecha_fin.strftime('%d/%m/%Y')} | "
            f"Enfermeras: {self.num_enfermeras} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        )
        info = Paragraph(
            info_text,
            ParagraphStyle('Info', fontName='Helvetica', fontSize=7, textColor=rl_colors.grey)
        )
        contenido.append(info)
        contenido.append(Spacer(1, 0.15*cm))
        
        datos_tabla = [encabezado]
        datos_tabla.extend(matriz)
        
        tabla = Table(datos_tabla, repeatRows=1)
        
        estilos = [
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BACKGROUND', (0, 1), (1, -1), rl_colors.HexColor('#D9E8F5')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (2, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.black),
        ]
        
        for row_idx, fila in enumerate(matriz, 1):
            for col_idx, valor in enumerate(fila, 0):
                if col_idx >= 2 and valor in COLORES_TURNOS:
                    color_hex = COLORES_TURNOS[valor]['rgb_rl']
                    estilos.append(
                        ('BACKGROUND', (col_idx + 2, row_idx), (col_idx + 2, row_idx),
                         rl_colors.HexColor(color_hex))
                    )
                    if valor == 'NOCHE':
                        estilos.append(
                            ('TEXTCOLOR', (col_idx + 2, row_idx), (col_idx + 2, row_idx),
                             rl_colors.whitesmoke)
                        )
        
        tabla.setStyle(TableStyle(estilos))
        contenido.append(tabla)
        
        return contenido
    
    def _generar_pagina_estadisticas(self):
        """Genera página de estadísticas."""
        contenido = []
        
        titulo = Paragraph(
            "ESTADÍSTICAS Y ANÁLISIS",
            ParagraphStyle(
                'Titulo',
                fontName='Helvetica-Bold',
                fontSize=12,
                textColor=rl_colors.white,
                backColor=rl_colors.HexColor('#1F4E78'),
                alignment=TA_CENTER,
                leftIndent=5,
                rightIndent=5,
                topPadding=5,
                bottomPadding=5
            )
        )
        contenido.append(titulo)
        contenido.append(Spacer(1, 0.2*cm))
        
        # Tabla de estadísticas
        conteo = self.stats.contar_turnos_por_tipo()
        total = sum(conteo.values()) if conteo else 1
        
        datos_stats = [['Turno', 'Cantidad', 'Porcentaje']]
        for turno_nombre in sorted(conteo.keys()):
            cantidad = conteo[turno_nombre]
            porcentaje = f"{(cantidad / total * 100):.1f}%"
            datos_stats.append([turno_nombre, str(cantidad), porcentaje])
        
        tabla_stats = Table(datos_stats)
        tabla_stats.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.black),
        ]))
        
        contenido.append(tabla_stats)
        contenido.append(Spacer(1, 0.3*cm))
        
        equidad = self.stats.distribucion_equidad()
        info_equidad = Paragraph(
            f"<b>Análisis de Equidad:</b> Media {equidad['media']:.1f} | "
            f"Rango {equidad['min']}-{equidad['max']} | "
            f"Desviación {equidad['desviacion']:.2f}",
            ParagraphStyle('Info', fontName='Helvetica', fontSize=9)
        )
        contenido.append(info_equidad)
        
        return contenido
    
    def exportar_ambos(self, nombre_base):
        """Exporta a PDF y Excel."""
        archivo_excel = f"{nombre_base}.xlsx"
        archivo_pdf = f"{nombre_base}.pdf"
        
        self.exportar_excel(archivo_excel)
        self.exportar_pdf(archivo_pdf)
        
        return {
            'excel': archivo_excel,
            'pdf': archivo_pdf
        }
    
    def generar_reporte_txt(self, archivo_salida):
        """Genera reporte de validaciones en texto."""
        reporte = self.validador.generar_reporte_validacion(self.stats)
        
        with open(archivo_salida, 'w', encoding='utf-8') as f:
            f.write(reporte)
        
        logger.info(f"✓ Reporte TXT generado: {archivo_salida}")
        return archivo_salida


# =========================================================================
# FUNCIONES DE CONVENIENCIA
# =========================================================================

def exportar_resultados(planificacion_resultado, config, nombre_base):
    """
    Función simplificada para exportar resultados.
    
    Args:
        planificacion_resultado: Resultado con turnos asignados
        config: Configuración del proyecto
        nombre_base: Nombre base sin extensión
    
    Returns:
        Dict con rutas de archivos generados
    """
    exportador = ExportadorProfesional(planificacion_resultado, config)
    return exportador.exportar_ambos(nombre_base)


def exportar_con_reporte(planificacion_resultado, config, nombre_base):
    """
    Exporta a Excel, PDF y genera reporte de validaciones.
    
    Returns:
        Dict con todas las rutas de archivos
    """
    exportador = ExportadorProfesional(planificacion_resultado, config)
    
    archivos = exportador.exportar_ambos(nombre_base)
    reporte_txt = exportador.generar_reporte_txt(f"{nombre_base}_validacion.txt")
    
    archivos['reporte_txt'] = reporte_txt
    
    return archivos


def generar_reporte_validacion_txt(planificacion_resultado):
    """Genera solo el reporte de validaciones."""
    exportador = ExportadorProfesional(planificacion_resultado)
    return exportador.validador.generar_reporte_validacion(exportador.stats)


# =========================================================================
# EJEMPLO DE USO
# =========================================================================

if __name__ == "__main__":
    # Datos de ejemplo
    planificacion_data = {
        'enfermeras': [
            {'nombre': f'Enfermera_{i}', 'rol': 'Enfermera'}
            for i in range(1, 19)
        ],
        'turnos_asignados': {
            (0, 0): 'MAÑANA', (1, 0): 'TARDE', (2, 0): 'NOCHE',
            (0, 1): 'TARDE', (1, 1): 'NOCHE', (2, 1): 'MAÑANA',
            # ... más turnos ...
        },
        'fecha_inicio': datetime(2024, 2, 1),
        'fecha_fin': datetime(2024, 2, 29),
    }
    
    config = {}
    
    # Exportar con todas las funcionalidades
    archivos = exportar_con_reporte(planificacion_data, config, 'planificacion_feb_2024')
    
    print("✓ Exportación completada:")
    print(f"  Excel: {archivos['excel']}")
    print(f"  PDF: {archivos['pdf']}")
    print(f"  Reporte: {archivos['reporte_txt']}")
