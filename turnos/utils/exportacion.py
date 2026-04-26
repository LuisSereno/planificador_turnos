# =========================================================================
# EXPORTACION.PY - VERSIÓN COMPLETA CON 7 HOJAS
# =========================================================================
"""
Versión definitiva con 7 hojas:
1. Planilla Vertical
2. Planilla Horizontal (como la web)
3. Estadísticas
4. Por Enfermera
5. Cobertura
6. Equidad
7. Validaciones
"""

from io import BytesIO
from datetime import datetime, timedelta
import json
import csv
from collections import defaultdict
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from icalendar import Calendar, Event

    ICAL_AVAILABLE = True
except ImportError:
    ICAL_AVAILABLE = False

import logging

logger = logging.getLogger(__name__)

# --- Configuración de Estilos (CORREGIDO: Sin tilde) ---
COLORES_TURNOS = {
    'MANANA': {'rgb': 'FFC107', 'rgb_rl': '#FFC107', 'nombre': 'Mañana'},
    'TARDE': {'rgb': '00BCD4', 'rgb_rl': '#00BCD4', 'nombre': 'Tarde'},
    'NOCHE': {'rgb': '424242', 'rgb_rl': '#424242', 'nombre': 'Noche'},
    'LIBRE': {'rgb': 'E0E0E0', 'rgb_rl': '#E0E0E0', 'nombre': 'Libre'},
}

COLOR_ENCABEZADO = "1F4E78"
COLOR_EXITO = "6BCB77"
COLOR_ALERTA = "FFD93D"

BORDE_DELGADO = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =========================================================================
# FUNCIONES TRADUCTORAS
# =========================================================================

def _traducir_modelo_a_diccionario_VERTICAL(ejecucion):
    """Convierte las asignaciones a formato VERTICAL"""
    logger.info(f"🔵 INICIO _traducir_modelo_a_diccionario_VERTICAL para ejecución {ejecucion.id}")

    if not hasattr(ejecucion, 'planilla_generada') or not ejecucion.planilla_generada:
        logger.warning(f"⚠️  No se encontró planilla para ejecución {ejecucion.id}")
        return {}

    asignaciones = ejecucion.planilla_generada.asignaciones.select_related('turno', 'enfermera').all()
    logger.info(f"📊 Total de asignaciones encontradas: {asignaciones.count()}")

    planilla_dict = defaultdict(lambda: defaultdict(list))
    contador_por_turno = defaultdict(int)

    for asignacion in asignaciones:
        if asignacion.turno:
            dia_key = f"dia_{(asignacion.fecha - ejecucion.configuracion.fecha_inicio).days + 1}"

            if not asignacion.es_dia_libre:
                planilla_dict[dia_key][asignacion.turno.nombre].append(asignacion.enfermera.nombre)
                contador_por_turno[asignacion.turno.nombre] += 1

    logger.info(f"📈 RESUMEN DE ASIGNACIONES POR TURNO:")
    for turno, cantidad in contador_por_turno.items():
        logger.info(f"   - {turno}: {cantidad} asignaciones")

    return dict(planilla_dict)


def _traducir_modelo_a_diccionario_HORIZONTAL(ejecucion):
    """Convierte las asignaciones a formato HORIZONTAL"""
    logger.info(f"🟢 INICIO _traducir_modelo_a_diccionario_HORIZONTAL para ejecución {ejecucion.id}")

    if not hasattr(ejecucion, 'planilla_generada') or not ejecucion.planilla_generada:
        return {}, []

    asignaciones = ejecucion.planilla_generada.asignaciones.select_related('turno', 'enfermera').all()

    datos_horizontales = defaultdict(dict)
    fechas_del_periodo = set()

    for asignacion in asignaciones:
        fechas_del_periodo.add(asignacion.fecha)

        # Corregido: LIBRE cuando es_dia_libre O cuando turno es null
        if asignacion.es_dia_libre or not asignacion.turno:
            datos_horizontales[asignacion.enfermera.id][asignacion.fecha] = 'LIBRE'
        else:
            datos_horizontales[asignacion.enfermera.id][asignacion.fecha] = asignacion.turno.nombre

    return dict(datos_horizontales), sorted(list(fechas_del_periodo))


# =========================================================================
# FUNCIÓN: GENERAR EXCEL CON 7 HOJAS
# =========================================================================

def generar_excel_planilla(ejecucion):
    """Genera Excel con 7 hojas profesionales"""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl no está instalado")

    try:
        logger.info(f"🚀 INICIO generar_excel_planilla para ejecución {ejecucion.id}")

        wb = Workbook()
        fecha_inicio = ejecucion.configuracion.fecha_inicio

        # ===== HOJA 1: PLANILLA VERTICAL =====
        logger.info("📄 Generando Hoja 1: Planilla Vertical")

        ws_vertical = wb.active
        ws_vertical.title = "Planilla Vertical"

        header_fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        ws_vertical['A1'] = f"Planificación: {ejecucion.configuracion.nombre}"
        ws_vertical['A1'].font = Font(size=14, bold=True)
        ws_vertical.merge_cells('A1:D1')

        current_row = 5
        headers = ['Día', 'Fecha', 'Turno', 'Enfermeras']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_vertical.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDE_DELGADO

        current_row += 1

        planilla_vertical = _traducir_modelo_a_diccionario_VERTICAL(ejecucion)

        for i in range(1, ejecucion.configuracion.num_dias + 1):
            dia_key = f"dia_{i}"
            fecha_actual = fecha_inicio + timedelta(days=i - 1)
            turnos = planilla_vertical.get(dia_key, {})

            for turno_tipo in ['MANANA', 'TARDE', 'NOCHE']:
                enfermeras = turnos.get(turno_tipo, [])
                turno_display = 'MAÑANA' if turno_tipo == 'MANANA' else turno_tipo

                ws_vertical.cell(row=current_row, column=1).value = i
                ws_vertical.cell(row=current_row, column=2).value = fecha_actual.strftime('%d/%m/%Y')
                ws_vertical.cell(row=current_row, column=3).value = turno_display
                ws_vertical.cell(row=current_row, column=4).value = ', '.join(
                    enfermeras) if enfermeras else 'Sin asignar'

                if turno_tipo in COLORES_TURNOS:
                    ws_vertical.cell(row=current_row, column=3).fill = PatternFill(
                        start_color=COLORES_TURNOS[turno_tipo]['rgb'],
                        end_color=COLORES_TURNOS[turno_tipo]['rgb'],
                        fill_type="solid"
                    )
                    if turno_tipo == 'NOCHE':
                        ws_vertical.cell(row=current_row, column=3).font = Font(color="FFFFFF", bold=True)

                for col_idx in range(1, 5):
                    ws_vertical.cell(row=current_row, column=col_idx).border = BORDE_DELGADO
                current_row += 1

        ws_vertical.column_dimensions['A'].width = 8
        ws_vertical.column_dimensions['B'].width = 15
        ws_vertical.column_dimensions['C'].width = 12
        ws_vertical.column_dimensions['D'].width = 50

        # ===== HOJA 2: PLANILLA HORIZONTAL =====
        logger.info("📄 Generando Hoja 2: Planilla Horizontal")

        ws_horizontal = wb.create_sheet("Planilla Horizontal")

        datos_horizontales, fechas = _traducir_modelo_a_diccionario_HORIZONTAL(ejecucion)

        ws_horizontal.cell(row=1, column=1).value = 'Enfermera'
        ws_horizontal.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
        ws_horizontal.cell(row=1, column=1).fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO,
                                                               fill_type="solid")

        for col_idx, fecha in enumerate(fechas, 2):
            cell = ws_horizontal.cell(row=1, column=col_idx)
            dia_semana = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom'][fecha.weekday()]
            cell.value = f"{fecha.strftime('%d/%m')}\n{dia_semana}"
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")

        for row_idx, enfermera in enumerate(ejecucion.configuracion.enfermeras.all(), 2):
            ws_horizontal.cell(row=row_idx, column=1).value = enfermera.nombre
            ws_horizontal.cell(row=row_idx, column=1).font = Font(bold=True)

            turnos_enfermera = datos_horizontales.get(enfermera.id, {})

            for col_idx, fecha in enumerate(fechas, 2):
                turno = turnos_enfermera.get(fecha, '-')
                turno_display = 'MAÑANA' if turno == 'MANANA' else turno

                cell = ws_horizontal.cell(row=row_idx, column=col_idx)
                cell.value = turno_display if turno_display != '-' else ''
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = BORDE_DELGADO

                turno_clave = 'MANANA' if turno == 'MANANA' else turno
                if turno_clave in COLORES_TURNOS:
                    fill_color = COLORES_TURNOS[turno_clave]['rgb']
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    if turno_clave == 'NOCHE':
                        cell.font = Font(color="FFFFFF", bold=True, size=9)
                    else:
                        cell.font = Font(bold=True, size=9)

        ws_horizontal.column_dimensions['A'].width = 20
        for col_idx in range(2, len(fechas) + 2):
            ws_horizontal.column_dimensions[get_column_letter(col_idx)].width = 10

        # ===== HOJA 3: ESTADÍSTICAS =====
        logger.info("📄 Generando Hoja 3: Estadísticas")

        ws_stats = wb.create_sheet("Estadísticas")
        ws_stats['A1'] = "Estadísticas de la Planificación"
        ws_stats['A1'].font = Font(size=12, bold=True, color="FFFFFF")
        ws_stats['A1'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws_stats.merge_cells('A1:C1')

        row = 3
        stats_data = [
            ("Penalización Total:", ejecucion.penalizacion_total or 0),
            ("Es Óptima:", "Sí" if ejecucion.es_optima else "No"),
            ("Duración (segundos):", ejecucion.duracion or 0),
            ("Estado:", ejecucion.estado),
        ]

        for label, value in stats_data:
            ws_stats.cell(row=row, column=1).value = label
            ws_stats.cell(row=row, column=2).value = value
            ws_stats.cell(row=row, column=1).font = Font(bold=True)
            row += 1

        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 20

        # ===== HOJA 4: POR ENFERMERA =====
        logger.info("📄 Generando Hoja 4: Por Enfermera")

        ws_enf = wb.create_sheet("Por Enfermera")
        ws_enf['A1'] = "Distribución de Turnos por Enfermera"
        ws_enf['A1'].font = Font(size=12, bold=True, color="FFFFFF")
        ws_enf['A1'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws_enf.merge_cells('A1:F1')

        row = 3
        headers_enf = ['Enfermera', 'Total Turnos', 'MAÑANA', 'TARDE', 'NOCHE', '% Ocupación']
        for col_idx, header in enumerate(headers_enf, 1):
            cell = ws_enf.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")

        row += 1
        for enfermera in ejecucion.configuracion.enfermeras.all():
            total_turnos = 0
            turnos_por_tipo = {'MANANA': 0, 'TARDE': 0, 'NOCHE': 0}

            for i in range(1, ejecucion.configuracion.num_dias + 1):
                dia_key = f"dia_{i}"
                dia_turnos = planilla_vertical.get(dia_key, {})
                for turno_tipo in ['MANANA', 'TARDE', 'NOCHE']:
                    if enfermera.nombre in dia_turnos.get(turno_tipo, []):
                        total_turnos += 1
                        turnos_por_tipo[turno_tipo] += 1

            ws_enf.cell(row=row, column=1).value = enfermera.nombre
            ws_enf.cell(row=row, column=2).value = total_turnos
            ws_enf.cell(row=row, column=3).value = turnos_por_tipo['MANANA']
            ws_enf.cell(row=row, column=4).value = turnos_por_tipo['TARDE']
            ws_enf.cell(row=row, column=5).value = turnos_por_tipo['NOCHE']

            ocupacion = (
                        total_turnos / ejecucion.configuracion.num_dias * 100) if ejecucion.configuracion.num_dias > 0 else 0
            ws_enf.cell(row=row, column=6).value = f"{ocupacion:.1f}%"

            row += 1

        for col in range(1, 7):
            ws_enf.column_dimensions[get_column_letter(col)].width = 16

        # ===== HOJA 5: COBERTURA =====
        logger.info("📄 Generando Hoja 5: Cobertura")

        ws_cob = wb.create_sheet("Cobertura")
        ws_cob['A1'] = "Análisis de Cobertura Diaria"
        ws_cob['A1'].font = Font(size=12, bold=True, color="FFFFFF")
        ws_cob['A1'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws_cob.merge_cells('A1:F1')

        row = 3
        headers_cob = ['Fecha', 'Día', 'MAÑANA', 'TARDE', 'NOCHE', 'Total']
        for col_idx, header in enumerate(headers_cob, 1):
            cell = ws_cob.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")

        row += 1
        for i in range(1, ejecucion.configuracion.num_dias + 1):
            dia_key = f"dia_{i}"
            fecha_actual = fecha_inicio + timedelta(days=i - 1)
            fecha_str = fecha_actual.strftime('%d/%m/%Y')
            dia_semana = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom'][fecha_actual.weekday()]

            ws_cob.cell(row=row, column=1).value = fecha_str
            ws_cob.cell(row=row, column=2).value = dia_semana

            dia_turnos = planilla_vertical.get(dia_key, {})
            total_dia = 0

            for col_idx, turno_tipo in enumerate(['MANANA', 'TARDE', 'NOCHE'], 3):
                cantidad = len(dia_turnos.get(turno_tipo, []))
                cell = ws_cob.cell(row=row, column=col_idx)
                cell.value = cantidad
                total_dia += cantidad

                if cantidad > 0 and turno_tipo in COLORES_TURNOS:
                    cell.fill = PatternFill(
                        start_color=COLORES_TURNOS[turno_tipo]['rgb'],
                        end_color=COLORES_TURNOS[turno_tipo]['rgb'],
                        fill_type="solid"
                    )

            ws_cob.cell(row=row, column=6).value = total_dia
            ws_cob.cell(row=row, column=6).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF",
                                                              fill_type="solid")
            ws_cob.cell(row=row, column=6).font = Font(bold=True)

            row += 1

        for col in range(1, 7):
            ws_cob.column_dimensions[get_column_letter(col)].width = 15

        # ===== HOJA 6: EQUIDAD =====
        logger.info("📄 Generando Hoja 6: Equidad")

        ws_equidad = wb.create_sheet("Equidad")
        ws_equidad['A1'] = "Análisis de Equidad"
        ws_equidad['A1'].font = Font(size=12, bold=True, color="FFFFFF")
        ws_equidad['A1'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws_equidad.merge_cells('A1:B1')

        row = 3
        turnos_por_enf = {}
        for enfermera in ejecucion.configuracion.enfermeras.all():
            total = 0
            for i in range(1, ejecucion.configuracion.num_dias + 1):
                dia_key = f"dia_{i}"
                for turno_tipo in ['MANANA', 'TARDE', 'NOCHE']:
                    if enfermera.nombre in planilla_vertical.get(dia_key, {}).get(turno_tipo, []):
                        total += 1
            turnos_por_enf[enfermera.nombre] = total

        if turnos_por_enf:
            valores = list(turnos_por_enf.values())
            media = sum(valores) / len(valores)
            minimo = min(valores)
            maximo = max(valores)
            diferencia = maximo - minimo
            varianza = sum((x - media) ** 2 for x in valores) / len(valores)
            desviacion = varianza ** 0.5

            datos_equidad = [
                ("Promedio de turnos:", f"{media:.1f}"),
                ("Mínimo:", f"{minimo}"),
                ("Máximo:", f"{maximo}"),
                ("Diferencia:", f"{diferencia}"),
                ("Desviación estándar:", f"{desviacion:.2f}"),
            ]

            for label, valor in datos_equidad:
                ws_equidad.cell(row=row, column=1).value = label
                ws_equidad.cell(row=row, column=2).value = valor
                ws_equidad.cell(row=row, column=1).font = Font(bold=True)
                row += 1

        ws_equidad.column_dimensions['A'].width = 30
        ws_equidad.column_dimensions['B'].width = 20

        # ===== HOJA 7: VALIDACIONES =====
        logger.info("📄 Generando Hoja 7: Validaciones")

        ws_val = wb.create_sheet("Validaciones")
        ws_val['A1'] = "Reporte de Validaciones"
        ws_val['A1'].font = Font(size=12, bold=True, color="FFFFFF")
        ws_val['A1'].fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
        ws_val.merge_cells('A1:B1')

        row = 3
        ws_val.cell(row=row, column=1).value = "Validación"
        ws_val.cell(row=row, column=2).value = "Resultado"
        for col in [1, 2]:
            ws_val.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws_val.cell(row=row, column=col).fill = PatternFill(start_color=COLOR_ENCABEZADO,
                                                                end_color=COLOR_ENCABEZADO, fill_type="solid")

        row += 1
        ws_val.cell(row=row, column=1).value = "Estado"
        ws_val.cell(row=row, column=2).value = ejecucion.estado
        ws_val.cell(row=row, column=2).fill = PatternFill(start_color=COLOR_EXITO, end_color=COLOR_EXITO,
                                                          fill_type="solid")

        row += 1
        ws_val.cell(row=row, column=1).value = "Es Óptima"
        ws_val.cell(row=row, column=2).value = "Sí" if ejecucion.es_optima else "No"
        color = COLOR_EXITO if ejecucion.es_optima else COLOR_ALERTA
        ws_val.cell(row=row, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws_val.column_dimensions['A'].width = 25
        ws_val.column_dimensions['B'].width = 30

        # Guardar
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f"✅ Excel generado con 7 hojas para ejecución {ejecucion.id}")
        return buffer

    except Exception as e:
        logger.error(f"❌ Error al generar Excel: {str(e)}", exc_info=True)
        raise


# =========================================================================
# RESTO DE FUNCIONES (PDF, CSV, JSON, ICAL) - Sin cambios
# =========================================================================

def generar_pdf_planilla(ejecucion):
    """Genera PDF"""
    if not PDF_AVAILABLE:
        raise ImportError("reportlab no está instalado")
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        story = []

        planilla = _traducir_modelo_a_diccionario_VERTICAL(ejecucion)
        fecha_inicio = ejecucion.configuracion.fecha_inicio
        styles = getSampleStyleSheet()

        title = Paragraph(f"Planificación: {ejecucion.configuracion.nombre}", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))

        data = [['Día', 'Fecha', 'Turno', 'Enfermeras']]
        for i in range(1, ejecucion.configuracion.num_dias + 1):
            dia_key = f"dia_{i}"
            fecha_actual = fecha_inicio + timedelta(days=i - 1)
            turnos = planilla.get(dia_key, {})
            for turno_tipo in ['MANANA', 'TARDE', 'NOCHE']:
                enfermeras = turnos.get(turno_tipo, [])
                turno_display = 'MAÑANA' if turno_tipo == 'MANANA' else turno_tipo
                data.append([str(i), fecha_actual.strftime('%d/%m/%Y'), turno_display, ', '.join(enfermeras)])

        table = Table(data, colWidths=[50, 100, 100, 400])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(table)

        doc.build(story)
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error al generar PDF: {str(e)}", exc_info=True)
        raise


def generar_csv_planilla(ejecucion):
    """Genera CSV"""
    try:
        buffer = BytesIO()
        text_buffer = io.TextIOWrapper(buffer, encoding='utf-8-sig', newline='')
        writer = csv.writer(text_buffer, delimiter=';')
        writer.writerow(['Día', 'Fecha', 'Turno', 'Enfermeras'])

        planilla = _traducir_modelo_a_diccionario_VERTICAL(ejecucion)
        fecha_inicio = ejecucion.configuracion.fecha_inicio

        for i in range(1, ejecucion.configuracion.num_dias + 1):
            dia_key = f"dia_{i}"
            fecha_actual = fecha_inicio + timedelta(days=i - 1)
            turnos = planilla.get(dia_key, {})
            for turno_tipo in ['MANANA', 'TARDE', 'NOCHE']:
                enfermeras = turnos.get(turno_tipo, [])
                turno_display = 'MAÑANA' if turno_tipo == 'MANANA' else turno_tipo
                writer.writerow([i, fecha_actual.strftime('%d/%m/%Y'), turno_display, ', '.join(enfermeras)])

        text_buffer.flush()
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error al generar CSV: {str(e)}", exc_info=True)
        raise


def generar_json_planilla(ejecucion):
    """Genera JSON"""
    try:
        planilla_serializable = _traducir_modelo_a_diccionario_VERTICAL(ejecucion)
        data = {
            'configuracion': {
                'id': ejecucion.configuracion.id,
                'nombre': ejecucion.configuracion.nombre,
            },
            'ejecucion': {
                'id': ejecucion.id,
                'estado': ejecucion.estado,
            },
            'planilla': planilla_serializable,
            'generado': datetime.now().isoformat(),
        }
        json_str = json.dumps(data, indent=2, ensure_ascii=False)
        buffer = BytesIO(json_str.encode('utf-8'))
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error al generar JSON: {str(e)}", exc_info=True)
        raise


def generar_ical_planilla(ejecucion):
    """Genera iCalendar"""
    if not ICAL_AVAILABLE:
        raise ImportError("icalendar no está instalado")
    try:
        cal = Calendar()
        cal.add('prodid', '-//Planificador de Turnos//ES')
        cal.add('version', '2.0')

        planilla = _traducir_modelo_a_diccionario_VERTICAL(ejecucion)
        fecha_inicio = ejecucion.configuracion.fecha_inicio
        horarios = {'MANANA': ('07:00', '15:00'), 'TARDE': ('15:00', '23:00'), 'NOCHE': ('23:00', '07:00')}

        for i in range(1, ejecucion.configuracion.num_dias + 1):
            dia_key = f"dia_{i}"
            fecha_actual = fecha_inicio + timedelta(days=i - 1)
            turnos = planilla.get(dia_key, {})
            for turno_tipo, (hora_inicio, hora_fin) in horarios.items():
                if enfermeras := turnos.get(turno_tipo, []):
                    event = Event()
                    turno_display = 'Mañana' if turno_tipo == 'MANANA' else turno_tipo.capitalize()
                    event.add('summary', f'Turno {turno_display}')
                    inicio_hora, inicio_min = map(int, hora_inicio.split(':'))
                    fin_hora, fin_min = map(int, hora_fin.split(':'))
                    dt_inicio = datetime.combine(fecha_actual,
                                                 datetime.min.time().replace(hour=inicio_hora, minute=inicio_min))
                    if fin_hora < inicio_hora:
                        dt_fin = datetime.combine(fecha_actual + timedelta(days=1),
                                                  datetime.min.time().replace(hour=fin_hora, minute=fin_min))
                    else:
                        dt_fin = datetime.combine(fecha_actual,
                                                  datetime.min.time().replace(hour=fin_hora, minute=fin_min))
                    event.add('dtstart', dt_inicio)
                    event.add('dtend', dt_fin)
                    event.add('description', 'Enfermeras: ' + ', '.join(enfermeras))
                    cal.add_component(event)

        buffer = BytesIO(cal.to_ical())
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error al generar iCal: {str(e)}", exc_info=True)
        raise


def exportar_enfermeras_excel(enfermeras_queryset):
    """Exporta enfermeras"""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl no está instalado")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Enfermeras"

        headers = ['ID', 'Nombre', 'Email', 'Teléfono', 'DNI', 'Fecha Alta', 'Activa']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")

        for row_num, enfermera in enumerate(enfermeras_queryset, 2):
            ws.cell(row=row_num, column=1, value=enfermera.id)
            ws.cell(row=row_num, column=2, value=enfermera.nombre)
            ws.cell(row=row_num, column=3, value=enfermera.email)
            ws.cell(row=row_num, column=4, value=enfermera.telefono or '')
            ws.cell(row=row_num, column=5, value=enfermera.dni or '')
            ws.cell(row=row_num, column=6,
                    value=enfermera.fecha_alta.strftime('%d/%m/%Y') if enfermera.fecha_alta else '')
            ws.cell(row=row_num, column=7, value='Sí' if enfermera.activa else 'No')

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error al exportar enfermeras: {str(e)}", exc_info=True)
        raise
