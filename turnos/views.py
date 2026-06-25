# -*- coding: utf-8 -*-
"""
Views for turnos app
"""
import logging
from datetime import date, time, datetime, timedelta
from collections import defaultdict
import json

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Avg, F, ExpressionWrapper, DurationField
from django.http import HttpResponse, FileResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import FormView, DetailView
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView,
    TemplateView, View
)
from formtools.wizard.views import SessionWizardView
from django.core.serializers.json import DjangoJSONEncoder

from .models import Workspace
from .utils import generar_json_planilla, generar_ical_planilla
from .utils.exportacion import exportar_enfermeras_excel, generar_csv_planilla, generar_pdf_planilla, \
    generar_excel_planilla
from .forms import (
    EnfermeraForm, TipoTurnoForm, ConfiguracionPlanificacionForm,
    EjecucionRapidaForm,
    ImportarEnfermerasForm,
    ConfiguracionWizardStep1Form,
    ConfiguracionWizardStep2DemandaForm,
    ConfiguracionWizardStep3DurasForm,
    ConfiguracionWizardStep4BlandasForm
)
from .mixins import (
    OwnerRequiredMixin,
    FormMessageMixin, PaginationMixin, SearchMixin, FilterMixin
)
from .models import ConfiguracionPlanificacion, Ejecucion
from .models import (
    Enfermera, TipoTurno, Planilla, AsignacionTurno, PatronTurnos
)

logger = logging.getLogger(__name__)

# ========== Dashboard ==========

class DashboardView(LoginRequiredMixin, TemplateView):
    """Vista principal del dashboard"""
    template_name = 'turnos/dashboard.html'

    def get_context_data(self, **kwargs):
        logger.info(
            f"DashboardView accessed by user {self.request.user.username} ({self.request.user.id})"
        )
        context = super().get_context_data(**kwargs)

        logger.debug("Fetching dashboard statistics...")
        # Estadísticas generales
        try:
            stats = {
                'total_configuraciones': ConfiguracionPlanificacion.objects.count(),
                'ejecuciones_exitosas': Ejecucion.objects.filter(estado='COMPLETADA').count(),
                'enfermeras_activas': Enfermera.objects.filter(activa=True).count(),
                'dias_planificados': AsignacionTurno.objects.values('fecha').distinct().count(),
            }
            logger.debug(f"Stats retrieved: {stats}")
        except Exception as e:
            logger.error(f"Error fetching dashboard statistics: {str(e)}", exc_info=True)
            stats = {}
            messages.error(self.request, "Could not load dashboard statistics")

        logger.debug("Fetching recent executions...")
        # Ejecuciones recientes (últimas 5)
        try:
            ejecuciones_recientes = Ejecucion.objects.select_related('configuracion').order_by('-fecha_inicio')[:5]
            logger.debug(f"Recent executions count: {len(ejecuciones_recientes)}")
        except Exception as e:
            logger.error(f"Error fetching recent executions: {str(e)}", exc_info=True)
            ejecuciones_recientes = []
            messages.error(self.request, "Could not load recent executions")

        # Actividad reciente (opcional - puedes comentar si no lo usas)
        actividad_reciente = []

        # Añadir al contexto
        context['stats'] = stats
        context['ejecuciones_recientes'] = ejecuciones_recientes
        context['actividad_reciente'] = actividad_reciente

        return context


# ========== Configuraciones ==========

class ConfiguracionListView(LoginRequiredMixin, SearchMixin, FilterMixin, PaginationMixin, ListView):
    """Lista de configuraciones"""
    model = ConfiguracionPlanificacion
    template_name = 'turnos/configuration_list.html'
    context_object_name = 'configuraciones'
    paginate_by = 12

    search_fields = ['nombre', 'descripcion']
    filter_fields = {
        'estado': 'activa',
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.select_related('creado_por').prefetch_related('enfermeras', 'turnos')

        # Filtro de estado
        estado = self.request.GET.get('estado')
        if estado == 'activa':
            queryset = queryset.filter(activa=True)
        elif estado == 'inactiva':
            queryset = queryset.filter(activa=False)

        # Ordenamiento
        orden = self.request.GET.get('orden', '-fecha_creacion')
        queryset = queryset.order_by(orden)

        return queryset


class ConfiguracionDetailView(LoginRequiredMixin, DetailView):
    """Detalle de una configuración"""
    model = ConfiguracionPlanificacion
    template_name = 'turnos/configuration_detail.html'
    context_object_name = 'configuracion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Log the full configuration details
        logger.info(f"Configuración completa: {json.dumps(self.object.__dict__, default=str, separators=(',', ':'), ensure_ascii=False)}")

        # Ejecuciones recientes de esta configuración
        context['ejecuciones_recientes'] = self.object.ejecuciones.order_by('-fecha_inicio')[:5]

        return context


class ConfiguracionCreateView(LoginRequiredMixin, FormMessageMixin, CreateView):
    """Crear nueva configuración"""
    model = ConfiguracionPlanificacion
    form_class = ConfiguracionPlanificacionForm
    template_name = 'turnos/configuration_form.html'
    success_message = 'Configuración creada con éxito.'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from turnos.models import TipoTurno
        tipos = TipoTurno.objects.filter(activo=True).values('id', 'nombre', 'codigo_corto', 'es_sustituto_libre', 'es_incidencia')
        context['tipos_turno_json'] = json.dumps(list(tipos), ensure_ascii=False)
        return context

    def form_valid(self, form):
        logger.info(
            f"Creating new configuration by user {self.request.user.username} ({self.request.user.id})"
        )
        logger.debug(f"Form data: {json.dumps(form.cleaned_data, default=str, separators=(',', ':'), ensure_ascii=False)}")

        form.instance.creado_por = self.request.user
        try:
            response = super().form_valid(form)

            # Procesar patrones JSON
            patrones_json = form.cleaned_data.get('patrones_turnos_json', '')
            if patrones_json:
                self._procesar_patrones(self.object, patrones_json)

            logger.info(
                f"Configuration created successfully. ID: {self.object.id}, "
                f"Name: {self.object.nombre}, Created by: {self.object.creado_por.username}"
            )
            return response
        except Exception as e:
            logger.error(
                f"Error creating configuration: {str(e)}",
                exc_info=True,
                extra={'user': self.request.user, 'form_data': form.cleaned_data}
            )
            messages.error(self.request, "Error creating configuration")
            raise

    def _procesar_patrones(self, config, patrones_json):
        """Procesa el JSON de patrones y los asigna a la configuración"""
        import json
        try:
            if isinstance(patrones_json, str):
                patrones_data = json.loads(patrones_json)
            else:
                patrones_data = patrones_json

            if not isinstance(patrones_data, list):
                logger.warning("patrones_turnos JSON no es una lista")
                return

            # Limpiar patrones previos
            config.patrones_turnos.clear()

            # Procesar cada patrón
            for patron_data in patrones_data:
                try:
                    tipo = patron_data.get('tipo')
                    es_restriccion_dura = patron_data.get('es_restriccion_dura', True)
                    peso_penalizacion = patron_data.get('peso_penalizacion', 100)
                    configuracion = patron_data.get('configuracion', {})

                    # Crear o actualizar patrón
                    patron, _ = PatronTurnos.objects.get_or_create(
                        tipo=tipo,
                        defaults={
                            'nombre': f"Patrón {tipo}",
                            'es_restriccion_dura': es_restriccion_dura,
                            'peso_penalizacion': peso_penalizacion,
                            'configuracion': configuracion,
                            'activo': True
                        }
                    )

                    # Actualizar si ya existe
                    patron.es_restriccion_dura = es_restriccion_dura
                    patron.peso_penalizacion = peso_penalizacion
                    patron.configuracion = configuracion
                    patron.save()

                    # Añadir a la configuración
                    config.patrones_turnos.add(patron)

                    logger.debug(f"Patrón {tipo} procesado correctamente")
                except Exception as e:
                    logger.warning(f"Error procesando patrón: {e}")

        except json.JSONDecodeError as e:
            logger.error(f"Error parseando JSON de patrones: {e}")
        except Exception as e:
            logger.error(f"Error en _procesar_patrones: {e}")


class ConfiguracionUpdateView(LoginRequiredMixin, OwnerRequiredMixin, FormMessageMixin, UpdateView):
    """Editar configuración existente"""
    model = ConfiguracionPlanificacion
    form_class = ConfiguracionPlanificacionForm
    template_name = 'turnos/configuration_form.html'
    success_message = 'Configuración actualizada con éxito.'
    owner_field = 'creado_por'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from turnos.models import TipoTurno
        tipos = TipoTurno.objects.filter(activo=True).values('id', 'nombre', 'codigo_corto', 'es_sustituto_libre', 'es_incidencia')
        context['tipos_turno_json'] = json.dumps(list(tipos), ensure_ascii=False)
        return context

    def form_valid(self, form):
        logger.info(
            f"Updating configuration ID {self.object.id} by user {self.request.user.username} "
            f"({self.request.user.id})"
        )
        logger.debug(f"Form changes: {json.dumps(form.changed_data, default=str, separators=(',', ':'), ensure_ascii=False)}")

        try:
            response = super().form_valid(form)
            logger.info(
                f"Configuration updated successfully. ID: {self.object.id}, "
                f"Changes: {form.changed_data}, Updated by: {self.request.user.username}"
            )
            return response
        except Exception as e:
            logger.error(
                f"Error updating configuration ID {self.object.id}: {str(e)}",
                exc_info=True,
                extra={'user': self.request.user, 'form_data': form.cleaned_data}
            )
            messages.error(self.request, "Error updating configuration")
            raise


class ConfiguracionDeleteView(LoginRequiredMixin, OwnerRequiredMixin, DeleteView):
    """Eliminar configuración"""
    model = ConfiguracionPlanificacion
    template_name = 'turnos/configuration_confirm_delete.html'
    success_url = reverse_lazy('turnos:config_lista')
    owner_field = 'creado_por'

    def delete(self, request, *args, **kwargs):
        config = self.get_object()
        logger.info(
            f"User {request.user.username} ({request.user.id}) attempting to delete "
            f"configuration ID {config.id} - '{config.nombre}'"
        )

        try:
            response = super().delete(request, *args, **kwargs)
            logger.info(
                f"Configuration ID {config.id} deleted successfully by "
                f"user {request.user.username} ({request.user.id})"
            )
            messages.success(request, 'Configuración eliminada con éxito.')
            return response
        except Exception as e:
            logger.error(
                f"Error deleting configuration ID {config.id}: {str(e)}",
                exc_info=True,
                extra={'user': request.user, 'config_id': config.id}
            )
            messages.error(request, 'Error al eliminar la configuración')
            raise


class ConfiguracionWizardView(LoginRequiredMixin, TemplateView):
    """Wizard para crear configuración paso a paso"""
    template_name = 'turnos/config/wizard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['enfermeras'] = Enfermera.objects.filter(activa=True)
        context['turnos'] = TipoTurno.objects.filter(activo=True)
        return context


class ConfiguracionDuplicarView(LoginRequiredMixin, View):
    """Duplicar una configuración existente"""

    def post(self, request, pk):
        config_original = get_object_or_404(ConfiguracionPlanificacion, pk=pk)

        # Crear copia
        # PRIORIDAD: patrones_turnos_json es la fuente activa principal
        # legacy: patrones_turnos (ManyToMany)
        config_nueva = ConfiguracionPlanificacion.objects.create(
            nombre=f"{config_original.nombre} (Copia)",
            descripcion=config_original.descripcion,
            activa=config_original.activa,
            num_dias=config_original.num_dias,
            fecha_inicio=timezone.now().date(),
            demanda_por_turno=config_original.demanda_por_turno,
            restricciones_duras=config_original.restricciones_duras,
            restricciones_blandas=config_original.restricciones_blandas,
            patrones_turnos_json=config_original.patrones_turnos_json,  # ✅ Copiar JSON activo
            num_trabajadores=config_original.num_trabajadores,
            tiempo_maximo_segundos=config_original.tiempo_maximo_segundos,
            seed=config_original.seed,
            creado_por=request.user
        )

        # Copiar relaciones ManyToMany (LEGACY)
        config_nueva.enfermeras.set(config_original.enfermeras.all())
        config_nueva.turnos.set(config_original.turnos.all())
        config_nueva.patrones_turnos.set(config_original.patrones_turnos.all())  # Legacy

        messages.success(request, 'Configuración duplicada con éxito.')
        return redirect('turnos:config_detalle', pk=config_nueva.pk)


# ========== Asistente de Configuración (Wizard) ==========

FORMS = [
    ("paso1", ConfiguracionWizardStep1Form),
    ("paso2", ConfiguracionWizardStep2DemandaForm),
    ("paso3", ConfiguracionWizardStep3DurasForm),
    ("paso4", ConfiguracionWizardStep4BlandasForm),
]

TEMPLATES = {
    "paso1": "turnos/wizard/paso1_basico.html",
    "paso2": "turnos/wizard/paso2_demanda.html",
    "paso3": "turnos/wizard/paso3_duras.html",
    "paso4": "turnos/wizard/paso4_blandas.html",
}

class ConfiguracionWizardViewStepByStep(LoginRequiredMixin, SessionWizardView):
    """Wizard para crear configuración paso a paso"""
    url_name = 'turnos:config_wizard_step'

    def get_template_names(self):
        return [TEMPLATES[self.steps.current]]

    def done(self, form_list, **kwargs):
        logger.info(
            f"User {self.request.user.username} ({self.request.user.id}) "
            "completing configuration wizard"
        )
        form_data = self.get_all_cleaned_data()
        logger.debug(f"Form data processed: {json.dumps(form_data, default=str, separators=(',', ':'), ensure_ascii=False)}")

        try:
            # Log each form step data
            for i, form in enumerate(form_list):
                logger.debug(
                    f"Step {i+1} form data: "
                    f"{json.dumps(form.cleaned_data, default=str, separators=(',', ':'), ensure_ascii=False) if hasattr(form, 'cleaned_data') else 'No cleaned data'}"
                )

            # Los campos JSON ya han sido validados y procesados por los formularios
            logger.info("Using pre-validated form data for JSON fields")

            with transaction.atomic():
                # 2. Crear instancia del modelo
                logger.info("Creating ConfiguracionPlanificacion instance...")
                try:
                    config = ConfiguracionPlanificacion.objects.create(
                        nombre=form_data['nombre'],
                        descripcion=form_data.get('descripcion', ''),
                        num_dias=form_data['num_dias'],
                        fecha_inicio=form_data['fecha_inicio'],
                        demanda_por_turno=form_data['demanda_por_turno'],
                        restricciones_duras=form_data['restricciones_duras'],
                        restricciones_blandas=form_data['restricciones_blandas'],
                        num_trabajadores=form_data.get('num_trabajadores', 4),
                        tiempo_maximo_segundos=form_data.get('tiempo_maximo_segundos', 60),
                        seed=form_data.get('seed'),
                        creado_por=self.request.user,
                        activa=True
                    )
                    logger.info(
                        f"Configuration created - ID: {config.id}, "
                        f"Name: {config.nombre}, "
                        f"Num días: {config.num_dias}, "
                        f"Trabajadores: {config.num_trabajadores}"
                    )
                except Exception as e:
                    logger.error(
                        "Error creating configuration instance",
                        exc_info=True,
                        extra={'user': self.request.user, 'form_data': form_data}
                    )
                    raise

                # 3. Añadir relaciones ManyToMany
                logger.info("Adding ManyToMany relationships...")
                try:
                    enfermeras_count = len(form_data['enfermeras'])
                    turnos_count = len(form_data['turnos'])

                    logger.debug(
                        f"Assigning {enfermeras_count} enfermeras and {turnos_count} turnos "
                        f"to configuration {config.id}"
                    )

                    config.enfermeras.set(form_data['enfermeras'])
                    config.turnos.set(form_data['turnos'])
                    
                    if 'patrones_turnos' in form_data:
                        config.patrones_turnos.set(form_data['patrones_turnos'])

                    logger.info(
                        f"Relationships assigned - Enfermeras: {enfermeras_count}, "
                        f"Turnos: {turnos_count}"
                    )
                except Exception as e:
                    logger.error(
                        "Error assigning ManyToMany relationships",
                        exc_info=True,
                        extra={'config_id': config.id}
                    )
                    raise

            messages.success(self.request, f'Configuración "{config.nombre}" creada con éxito.')
            logger.info(
                f"Wizard completed successfully for configuration {config.id} "
                f"by user {self.request.user.username}"
            )
            return redirect('turnos:config_detalle', pk=config.pk)

        except Exception as e:
            logger.exception(
                "Unexpected error saving configuration from wizard",
                extra={'user': self.request.user}
            )
            messages.error(
                self.request,
                f"Se produjo un error inesperado al guardar la configuración: {e}"
            )
            return redirect('turnos:config_wizard')


# ========== Ejecuciones ==========

class EjecucionListView(LoginRequiredMixin, SearchMixin, FilterMixin, PaginationMixin, ListView):
    """Lista de ejecuciones"""
    model = Ejecucion
    template_name = 'turnos/ejecucion_list.html'
    context_object_name = 'ejecuciones'
    paginate_by = 20

    search_fields = ['configuracion__nombre']

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.select_related('configuracion', 'planilla_generada')

        # Filtros
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)

        config_id = self.request.GET.get('proyecto_turnos')
        if config_id:
            queryset = queryset.filter(configuracion_id=config_id)

        return queryset.order_by('-fecha_inicio')


class EjecucionDetailView(LoginRequiredMixin, DetailView):
    """Vista de detalle de una ejecución"""
    model = Ejecucion
    template_name = 'turnos/ejecucion_detail.html'
    context_object_name = 'ejecucion'

    def get_context_data(self, **kwargs):
        logger.info(
            f"Showing execution detail for ID {self.object.id} "
            f"to user {self.request.user.username}"
        )
        context = super().get_context_data(**kwargs)
        ejecucion = self.object
        logger.debug(
            f"Execution details - Status: {ejecucion.estado}, "
            f"Config ID: {ejecucion.configuracion.id}, "
            f"Start: {ejecucion.fecha_inicio}, "
            f"Duration: {ejecucion.duracion}"
        )

        # Extraer datos de validación del campo `mensajes`
        validacion_resultado = {
            'validaciones_ok': 0,
            'violaciones': 0
        }
        if isinstance(ejecucion.mensajes, dict):
            validacion_resultado['validaciones_ok'] = len(ejecucion.mensajes.get('validaciones', []))
            validacion_resultado['violaciones'] = len(ejecucion.mensajes.get('violaciones', []))

        context['validacion_resultado'] = validacion_resultado

        # Si tiene planilla, calcular datos para visualización
        if hasattr(ejecucion, 'planilla_generada'):
            planilla = ejecucion.planilla_generada

            # Obtener todas las asignaciones
            asignaciones = planilla.asignaciones.select_related('enfermera', 'turno').order_by('fecha',
                                                                                               'enfermera__nombre')
            # Añadir datos clave al contexto
            context['num_asignaciones'] = asignaciones.count()
            context['num_dias_total'] = ejecucion.configuracion.num_dias

            # Agrupar por enfermera para la tabla
            enfermeras_turnos = {}
            for asignacion in asignaciones:
                enfermera_id = asignacion.enfermera.id
                if enfermera_id not in enfermeras_turnos:
                    enfermeras_turnos[enfermera_id] = {
                        'enfermera': asignacion.enfermera,
                        'turnos': []
                    }

                enfermeras_turnos[enfermera_id]['turnos'].append({
                    'fecha': asignacion.fecha,
                    'turno': asignacion.turno,
                    'es_libre': asignacion.es_dia_libre,
                    'turno_color': self._get_turno_color(asignacion),
                    'horario': f"{asignacion.turno.hora_inicio.strftime('%H:%M')}-{asignacion.turno.hora_fin.strftime('%H:%M')}" if asignacion.turno else '-',
                })

            # Lista de días únicos
            fechas_unicas = sorted(set(a.fecha for a in asignaciones))
            dias = [{'fecha': f, 'dia_semana': f.strftime('%a')} for f in fechas_unicas]

            # Distribución de turnos por tipo
            distribucion_turnos = {}

            for turno in ejecucion.configuracion.turnos.all():
                count = asignaciones.filter(turno=turno, es_dia_libre=False).count()
                distribucion_turnos[turno.nombre] = {
                    'turno': turno,
                    'cantidad': count,
                    'porcentaje': round((count / asignaciones.count() * 100), 1) if asignaciones.count() > 0 else 0
                }

            # Días libres
            dias_libres_count = asignaciones.filter(es_dia_libre=True).count()
            distribucion_turnos['LIBRE'] = {
                'turno': None,
                'cantidad': dias_libres_count,
                'porcentaje': round((dias_libres_count / asignaciones.count() * 100),
                                    1) if asignaciones.count() > 0 else 0
            }

            # Carga de trabajo por enfermera (para la tabla)
            carga_enfermeras = []
            for enfermera in ejecucion.configuracion.enfermeras.all():
                turnos_trabajados = asignaciones.filter(enfermera=enfermera, es_dia_libre=False).count()
                dias_libres = asignaciones.filter(enfermera=enfermera, es_dia_libre=True).count()
                total_dias = ejecucion.configuracion.num_dias
                porcentaje = round((turnos_trabajados / total_dias * 100), 1) if total_dias > 0 else 0

                carga_enfermeras.append({
                    'enfermera': enfermera,
                    'turnos_trabajados': turnos_trabajados,
                    'dias_libres': dias_libres,
                    'total_dias': total_dias,
                    'porcentaje': porcentaje
                })

            context['enfermeras_turnos'] = list(enfermeras_turnos.values())
            context['dias'] = dias
            context['distribucion_turnos'] = distribucion_turnos
            context['carga_enfermeras'] = carga_enfermeras
            context['dias_json'] = json.dumps([
                {
                    'fecha': dia['fecha'].isoformat() if isinstance(dia['fecha'], date) else dia['fecha'],
                    'dia_semana': dia['dia_semana']
                }
                for dia in context.get('dias', [])
            ], cls=DjangoJSONEncoder)

            context['enfermeras_turnos_json'] = json.dumps([
                {
                    'enfermera': {
                        'id': enf['enfermera'].id,
                        'nombre': enf['enfermera'].nombre
                    },
                    'turnos': [
                        {
                            'fecha': turno['fecha'].isoformat() if isinstance(turno['fecha'], date) else turno['fecha'],
                            'turno': {
                                'id': turno['turno'].id,
                                'nombre': turno['turno'].nombre,
                                'hora_inicio': str(turno['turno'].hora_inicio) if turno['turno'] else None,
                                'hora_fin': str(turno['turno'].hora_fin) if turno['turno'] else None
                            } if turno['turno'] else None,
                            'turno_color': turno['turno_color'],
                            'es_libre': turno['es_libre'],
                            'horario': turno.get('horario', '-')
                        }
                        for turno in enf['turnos']
                    ]
                }
                for enf in context.get('enfermeras_turnos', [])
            ], cls=DjangoJSONEncoder)

        # Construir datos para la pestaña de Análisis Detallado
        context['analisis'] = self._build_analisis_detallado(ejecucion)

        return context

    def _get_turno_color(self, asignacion):
        """Devuelve el color del turno para el badge"""
        if asignacion.es_dia_libre:
            return 'secondary'

        if asignacion.turno:
            nombre_lower = asignacion.turno.nombre.lower()
            if nombre_lower == 'mañana':
                return 'warning'
            elif nombre_lower == 'tarde':
                return 'info'
            elif nombre_lower == 'noche':
                return 'dark'

        return 'primary'

    def _build_analisis_detallado(self, ejecucion):
        """Construye los datos para la pestaña de Análisis Detallado."""
        from turnos.models import Enfermera

        analisis = {
            'estado': ejecucion.estado,
            'es_optima': ejecucion.es_optima,
            'penalizacion_total': ejecucion.penalizacion_total,
            'duracion': ejecucion.duracion,
            'violaciones_duras': [],
            'violaciones_blandas': [],
            'warnings': [],
            'balances': [],
            'recomendaciones': [],
            'causas_posibles': [],
            'resumen_ejecucion': {},
        }

        resultado = ejecucion.resultado or {}
        mensajes = ejecucion.mensajes or {}

        # ── Estado general ──────────────────────────────────────────────
        if ejecucion.estado == 'COMPLETADA':
            analisis['estado_icono'] = 'fa-check-circle'
            analisis['estado_color'] = 'success'
            analisis['estado_texto'] = 'Ejecución completada con éxito'
        elif ejecucion.estado == 'INVIABLE':
            analisis['estado_icono'] = 'fa-times-circle'
            analisis['estado_color'] = 'danger'
            analisis['estado_texto'] = 'El solver no pudo encontrar una solución factible'
        elif ejecucion.estado == 'ERROR':
            analisis['estado_icono'] = 'fa-exclamation-triangle'
            analisis['estado_color'] = 'danger'
            analisis['estado_texto'] = 'Error durante la ejecución'
        else:
            analisis['estado_icono'] = 'fa-clock'
            analisis['estado_color'] = 'secondary'
            analisis['estado_texto'] = f'Estado: {ejecucion.estado}'

        # ── Violaciones desde resultado (motor nuevo) ──────────────────
        violaciones_raw = resultado.get('violaciones', [])
        for v in violaciones_raw:
            if isinstance(v, dict):
                enfermera_id = v.get('enfermera_id')
                enfermera_nombre = ''
                if enfermera_id:
                    try:
                        enfermera_nombre = Enfermera.objects.get(pk=enfermera_id).nombre
                    except Enfermera.DoesNotExist:
                        enfermera_nombre = f'ID {enfermera_id}'
                analisis['violaciones_duras'].append({
                    'tipo': v.get('tipo', 'DESCONOCIDO'),
                    'descripcion': v.get('descripcion', ''),
                    'fecha': v.get('fecha', ''),
                    'enfermera_id': enfermera_id,
                    'enfermera_nombre': enfermera_nombre,
                })

        # ── Violaciones desde mensajes (motor legacy) ───────────────────
        if isinstance(mensajes, dict):
            for v in mensajes.get('violaciones', []):
                if isinstance(v, dict):
                    analisis['violaciones_duras'].append({
                        'tipo': v.get('nombre', v.get('tipo', 'DESCONOCIDO')),
                        'descripcion': v.get('detalles', v.get('descripcion', '')),
                        'fecha': v.get('fecha', ''),
                        'enfermera_id': v.get('enfermera_id'),
                        'enfermera_nombre': v.get('enfermera_nombre', ''),
                    })

        # ── Warnings ────────────────────────────────────────────────────
        warnings_raw = resultado.get('warnings', [])
        for w in warnings_raw:
            if isinstance(w, dict):
                analisis['warnings'].append({
                    'tipo': w.get('tipo', 'WARNING'),
                    'descripcion': w.get('descripcion', w.get('mensaje', '')),
                })
            elif isinstance(w, str):
                analisis['warnings'].append({'tipo': 'WARNING', 'descripcion': w})

        # ── Balances por enfermera ──────────────────────────────────────
        balances_raw = resultado.get('balances', {})
        if isinstance(balances_raw, dict):
            for enf_id_str, bal in balances_raw.items():
                if isinstance(bal, dict):
                    enfermera_id = int(enf_id_str) if enf_id_str.isdigit() else enf_id_str
                    enfermera_nombre = ''
                    try:
                        enfermera_nombre = Enfermera.objects.get(pk=enfermera_id).nombre
                    except (Enfermera.DoesNotExist, ValueError):
                        enfermera_nombre = f'Enfermera {enf_id_str}'
                    horas_asignadas = bal.get('horas_asignadas', 0)
                    horas_objetivo = bal.get('horas_objetivo', 0)
                    desviacion = bal.get('desviacion_horas', horas_asignadas - horas_objetivo)
                    analisis['balances'].append({
                        'enfermera_id': enfermera_id,
                        'enfermera_nombre': enfermera_nombre,
                        'horas_asignadas': round(horas_asignadas, 1),
                        'horas_objetivo': round(horas_objetivo, 1),
                        'desviacion_horas': round(desviacion, 1),
                        'desviacion_pct': round((desviacion / horas_objetivo * 100), 1) if horas_objetivo else 0,
                        'turnos_asignados': bal.get('turnos_asignados', 0),
                        'noches_asignadas': bal.get('noches_asignadas', 0),
                        'fines_semana_asignados': bal.get('fines_semana_asignados', 0),
                    })

        # ── Resumen de ejecución ────────────────────────────────────────
        config = ejecucion.configuracion
        analisis['resumen_ejecucion'] = {
            'nombre_config': config.nombre,
            'fecha_inicio': config.fecha_inicio.strftime('%d/%m/%Y') if config.fecha_inicio else '-',
            'num_dias': config.num_dias,
            'num_enfermeras': config.enfermeras.count(),
            'num_turnos': config.turnos.count(),
            'num_restricciones_duras': len(config.restricciones_duras or []),
            'num_restricciones_blandas': len(config.restricciones_blandas or []),
            'num_patrones': len(config.patrones_turnos_json or []),
            'horas_semanales': getattr(config, 'horas_semanales', 40),
            'horas_mensuales': getattr(config, 'horas_mensuales', 160),
            'horas_anuales': getattr(config, 'horas_anuales', 1800),
        }

        # ── Causas posibles ─────────────────────────────────────────────
        if ejecucion.estado == 'INVIABLE':
            analisis['causas_posibles'] = [
                'La demanda de personal es demasiado alta para el número de enfermeras disponibles.',
                'Las restricciones duras son incompatibles entre sí (ej: cobertura mínima alta + límite bajo de turnos consecutivos).',
                'No hay suficientes enfermeras para cubrir los turnos mínimos requeridos en todas las fechas.',
                'Las incidencias (vacaciones, bajas) bloquean demasiadas celdas, reduciendo la capacidad de maniobra del solver.',
            ]
        elif ejecucion.estado == 'ERROR':
            error_msg = resultado.get('error', '')
            analisis['causas_posibles'] = [
                f'Error técnico: {error_msg}' if error_msg else 'Error inesperado durante la ejecución.',
                'Revisa la configuración de la planificación y asegúrate de que todos los datos son correctos.',
                'Puede ser un problema de timeout: prueba a aumentar el tiempo máximo en la configuración avanzada.',
            ]

        # ── Recomendaciones ─────────────────────────────────────────────
        if ejecucion.estado == 'INVIABLE':
            analisis['recomendaciones'] = [
                'Reduce la demanda mínima de personal por turno (ej: de 2 a 1 enfermera).',
                'Aumenta el número de enfermeras seleccionadas en la configuración.',
                'Relaja las restricciones duras: aumenta el máximo de turnos consecutivos o reduce el descanso mínimo.',
                'Revisa las incidencias: si hay muchas vacaciones o bajas en el mismo período, considera reprogramar.',
                'Aumenta el tiempo máximo de resolución en la configuración avanzada.',
            ]
        elif ejecucion.estado == 'ERROR':
            analisis['recomendaciones'] = [
                'Revisa que todas las enfermeras seleccionadas estén activas.',
                'Verifica que los tipos de turno seleccionados tengan hora de inicio y fin configuradas.',
                'Aumenta el tiempo máximo de resolución (actualmente: ' + str(config.tiempo_maximo_segundos) + 's).',
                'Si el error persiste, revisa los logs del servidor para más detalles.',
            ]
        elif ejecucion.estado == 'COMPLETADA' and not ejecucion.es_optima:
            analisis['recomendaciones'] = [
                'La solución encontrada es factible pero no óptima. El solver se quedó sin tiempo antes de encontrar la mejor solución.',
                'Aumenta el tiempo máximo de resolución para permitir al solver explorar más combinaciones.',
                'Relaja algunas restricciones blandas (reduce su peso) para que el solver tenga más flexibilidad.',
                'Considera aumentar el número de trabajadores paralelos en la configuración avanzada.',
            ]
        elif ejecucion.estado == 'COMPLETADA' and ejecucion.es_optima:
            analisis['recomendaciones'] = [
                'La solución es óptima. No se requieren cambios para mejorar la planificación.',
                'Si deseas explorar alternativas, puedes modificar restricciones blandas y re-ejecutar.',
            ]

        # Penalizaciones
        if analisis['balances']:
            balances_con_desviacion = [b for b in analisis['balances'] if abs(b['desviacion_horas']) > 8]
            if balances_con_desviacion:
                nombres = ', '.join(b['enfermera_nombre'] for b in balances_con_desviacion[:5])
                analisis['recomendaciones'].append(
                    f'Enfermeras con desviación significativa de horas (>8h): {nombres}. '
                    f'Considera revisar sus contratos o ajustar las horas semanales de la configuración.'
                )

        return analisis


class EjecucionDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar ejecución"""
    model = Ejecucion
    template_name = 'turnos/ejecucion_confirm_delete.html'
    success_url = reverse_lazy('turnos:ejecucion_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Ejecución eliminada con éxito.')
        return super().delete(request, *args, **kwargs)


from .tasks import ejecutar_planificacion_motor_async  # NUEVO MOTOR - legacy task deprecated


# En turnos/views.py - Encontrar EjecutarPlanificacionView y reemplazarla:

class EjecutarPlanificacionView(LoginRequiredMixin, DetailView):
    """Vista para ejecutar una planificación"""
    model = ConfiguracionPlanificacion
    template_name = 'turnos/ejecutar_planificacion.html'
    context_object_name = 'configuracion'

    def get_context_data(self, **kwargs):
        logger.info(
            f"Preparing execution view for config ID {self.object.id} "
            f"by user {self.request.user.username}"
        )
        context = super().get_context_data(**kwargs)
        config = self.get_object()

        logger.debug("Validating configuration for execution...")
        # Validaciones
        errores = []
        enfermeras_count = config.enfermeras.count()
        turnos_count = config.turnos.count()

        if enfermeras_count < 2:
            error_msg = f'Se necesitan al menos 2 enfermeras (actual: {enfermeras_count})'
            logger.warning(error_msg)
            errores.append(error_msg)
        if turnos_count < 1:
            error_msg = f'Se necesita al menos 1 turno (actual: {turnos_count})'
            logger.warning(error_msg)
            errores.append(error_msg)

        context['errores'] = errores
        context['puede_ejecutar'] = len(errores) == 0

        logger.info(
            f"Validation results - Errors: {len(errores)}, "
            f"Can execute: {len(errores) == 0}"
        )

        return context

    def post(self, request, *args, **kwargs):
        """Ejecuta la planificación"""
        config_id = self.kwargs.get('pk')
        logger.info(
            f"Execution requested for config ID {config_id} "
            f"by user {request.user.username} ({request.user.id})"
        )

        try:
            config = ConfiguracionPlanificacion.objects.get(pk=config_id)
            logger.debug(
                f"Config found - ID: {config.id}, "
                f"Name: {config.nombre}, "
                f"Enfermeras: {config.enfermeras.count()}, "
                f"Turnos: {config.turnos.count()}"
            )
        except ConfiguracionPlanificacion.DoesNotExist:
            logger.error(f"Configuration ID {config_id} not found")
            messages.error(request, 'Configuración no encontrada')
            return redirect('turnos:config_lista')

        try:
            logger.info("Creating execution record...")
            with transaction.atomic():
                ejecucion = Ejecucion.objects.create(
                    configuracion=config,
                    estado='PENDIENTE'
                )
                logger.info(f"Execution record created - ID: {ejecucion.id}")

            logger.debug("Preparing Celery task...")
            try:
                config_id_int = int(config_id)
                logger.info(f"Dispatching Celery task for config ID {config_id_int}")

                task = ejecutar_planificacion_motor_async.delay(config_id_int)

                logger.info(
                    f"Celery task dispatched - Task ID: {task.id}, "
                    f"Execution ID: {ejecucion.id}"
                )

                messages.success(
                    request,
                    f'✓ Planificación enviada. Ejecución #{ejecucion.id}. Task: {task.id}'
                )
            except Exception as celery_error:
                logger.exception("Celery task dispatch failed", exc_info=True)
                ejecucion.estado = 'ERROR'
                ejecucion.mensajes = {'error': str(celery_error)}
                ejecucion.save()
                logger.error(
                    f"Execution marked as ERROR - ID: {ejecucion.id}, "
                    f"Message: {celery_error}"
                )
                messages.error(request, f'Error: {celery_error}')

            return redirect('turnos:ejecucion_detalle', pk=ejecucion.pk)

        except Exception as e:
            logger.exception(
                "Unexpected error during execution setup",
                exc_info=True,
                extra={'config_id': config_id}
            )
            messages.error(
                request,
                f'Error inesperado: {e}'
            )
            return redirect('turnos:config_detalle', pk=config_id)


class EjecucionRapidaView(LoginRequiredMixin, FormView):
    """Vista para ejecución rápida"""
    template_name = 'turnos/ejecutar_rapido.html'
    form_class = EjecucionRapidaForm
    success_url = reverse_lazy('turnos:ejecucion_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['enfermeras'] = Enfermera.objects.filter(activa=True)
        return context

    def form_valid(self, form):
        # Crear configuración rápida
        # ... implementación
        messages.success(self.request, 'Ejecución rápida iniciada.')
        return super().form_valid(form)


# ========== Enfermeras ==========

class EnfermeraListView(LoginRequiredMixin, SearchMixin, PaginationMixin, ListView):
    """Lista de enfermeras"""
    model = Enfermera
    template_name = 'turnos/enfermera_list.html'
    context_object_name = 'enfermeras'
    paginate_by = 15

    search_fields = ['nombre', 'email', 'dni']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtro de estado
        estado = self.request.GET.get('estado')
        if estado == 'activa':
            queryset = queryset.filter(activa=True)
        elif estado == 'inactiva':
            queryset = queryset.filter(activa=False)

        # Ordenamiento
        orden = self.request.GET.get('orden', 'nombre')
        queryset = queryset.order_by(orden)

        return queryset


class EnfermeraDetailView(LoginRequiredMixin, DetailView):
    """Detalle de una enfermera"""
    model = Enfermera
    template_name = 'turnos/enfermera_detail.html'
    context_object_name = 'enfermera'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas
        context['stats'] = {
            'total_turnos': AsignacionTurno.objects.filter(enfermera=self.object).count(),
            'configuraciones': ConfiguracionPlanificacion.objects.filter(enfermeras=self.object).count(),
            'turnos_noche': AsignacionTurno.objects.filter(
                enfermera=self.object,
                turno__nombre='NOCHE'
            ).count()
        }

        # Asignaciones recientes
        asignaciones = AsignacionTurno.objects.filter(
            enfermera=self.object
        ).select_related('turno', 'planilla__ejecucion').order_by('-fecha')[:10]

        asignaciones_recientes = []
        for a in asignaciones:
            asignaciones_recientes.append({
                'configuracion': a.planilla.ejecucion.configuracion,
                'fecha': a.fecha,
                'total_turnos': a.planilla.asignaciones.filter(fecha=a.fecha).count(),
                'ejecucion_id': a.planilla.ejecucion.id
            })

        context['asignaciones_recientes'] = asignaciones_recientes

        return context


class EnfermeraCreateView(LoginRequiredMixin, FormMessageMixin, CreateView):
    """Crear nueva enfermera"""
    model = Enfermera
    form_class = EnfermeraForm
    template_name = 'turnos/enfermera_form.html'
    success_message = 'Enfermera creada con éxito.'

    def form_valid(self, form):
        logger.info(
            f"User {self.request.user.username} ({self.request.user.id}) "
            "creating new enfermera"
        )
        logger.debug(f"Enfermera form data: {form.cleaned_data}")

        try:
            response = super().form_valid(form)
            logger.info(
                f"Enfermera created successfully - ID: {self.object.id}, "
                f"Name: {self.object.nombre}, Email: {self.object.email}"
            )
            return response
        except Exception as e:
            logger.error(
                f"Error creating enfermera: {str(e)}",
                exc_info=True,
                extra={'user': self.request.user, 'form_data': form.cleaned_data}
            )
            messages.error(self.request, "Error al crear la enfermera")
            raise


class EnfermeraUpdateView(LoginRequiredMixin, FormMessageMixin, UpdateView):
    """Editar enfermera"""
    model = Enfermera
    form_class = EnfermeraForm
    template_name = 'turnos/enfermera_form.html'
    success_message = 'Enfermera actualizada con éxito.'

    def form_valid(self, form):
        logger.info(
            f"User {self.request.user.username} ({self.request.user.id}) "
            f"updating enfermera ID {self.object.id}"
        )
        logger.debug(f"Changes: {form.changed_data}")

        try:
            response = super().form_valid(form)
            logger.info(
                f"Enfermera ID {self.object.id} updated successfully. "
                f"Changes: {form.changed_data}"
            )
            return response
        except Exception as e:
            logger.error(
                f"Error updating enfermera ID {self.object.id}: {str(e)}",
                exc_info=True,
                extra={'user': self.request.user, 'form_data': form.cleaned_data}
            )
            messages.error(self.request, "Error al actualizar la enfermera")
            raise


class EnfermeraDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar enfermera"""
    model = Enfermera
    template_name = 'turnos/enfermera_confirm_delete.html'
    success_url = reverse_lazy('turnos:enfermera_lista')

    def delete(self, request, *args, **kwargs):
        enfermera = self.get_object()
        logger.info(
            f"User {request.user.username} ({request.user.id}) attempting to delete "
            f"enfermera ID {enfermera.id} - {enfermera.nombre}"
        )

        try:
            response = super().delete(request, *args, **kwargs)
            logger.info(
                f"Enfermera ID {enfermera.id} deleted successfully by "
                f"user {request.user.username}"
            )
            messages.success(request, 'Enfermera eliminada con éxito.')
            return response
        except Exception as e:
            logger.error(
                f"Error deleting enfermera ID {enfermera.id}: {str(e)}",
                exc_info=True,
                extra={'user': request.user, 'enfermera_id': enfermera.id}
            )
            messages.error(request, 'Error al eliminar la enfermera')
            raise


class ImportarEnfermerasView(LoginRequiredMixin, FormView):
    """Importar enfermeras desde Excel"""
    template_name = 'turnos/enfermera_import.html'
    form_class = ImportarEnfermerasForm
    success_url = reverse_lazy('turnos:enfermera_lista')

    def form_valid(self, form):
        logger.info(
            f"User {self.request.user.username} ({self.request.user.id}) "
            "starting enfermeras import"
        )
        logger.debug(f"Import options: {form.cleaned_data}")

        archivo = form.cleaned_data['archivo']
        sobrescribir = form.cleaned_data['sobrescribir']

        try:
            import openpyxl
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            logger.info(f"Excel file loaded: {archivo.name}")

            total_rows = len(list(ws.iter_rows(min_row=2)))
            logger.info(f"Processing {total_rows} rows from Excel file")

            creadas = 0
            actualizadas = 0
            errores = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                nombre, email, telefono, dni, activa = row[:5]

                if not nombre or not email:
                    logger.warning(f"Skipping row with missing name/email: {row}")
                    continue

                activa = activa in ['Sí', 'Si', 'SI', 'sí', 'si', True, 1]
                logger.debug(f"Processing row: name={nombre}, email={email}, active={activa}")

                try:
                    enfermera_existente = Enfermera.objects.filter(email=email).first()

                    if enfermera_existente:
                        if sobrescribir:
                            logger.debug(f"Updating existing enfermera: {enfermera_existente.id}")
                            enfermera_existente.nombre = nombre
                            enfermera_existente.telefono = telefono or ''
                            enfermera_existente.dni = dni or ''
                            enfermera_existente.activa = activa
                            enfermera_existente.save()
                            actualizadas += 1
                            logger.debug(f"Enfermera {enfermera_existente.id} updated")
                    else:
                        logger.debug("Creating new enfermera")
                        Enfermera.objects.create(
                            nombre=nombre,
                            email=email,
                            telefono=telefono or '',
                            dni=dni or '',
                            activa=activa
                        )
                        creadas += 1
                        logger.debug(f"New enfermera created for email {email}")
                except Exception as e:
                    logger.error(
                        f"Error processing row {row}: {str(e)}",
                        exc_info=True,
                        extra={'row_data': row}
                    )
                    errores.append(str(e))

            logger.info(
                f"Import completed: {creadas} created, {actualizadas} updated, "
                f"{len(errores)} errors"
            )
            messages.success(
                self.request,
                f'Importación completada: {creadas} enfermeras creadas, {actualizadas} actualizadas.'
            )

        except Exception as e:
            logger.error(
                f"Error during import: {str(e)}",
                exc_info=True,
                extra={'user': self.request.user, 'file_name': archivo.name}
            )
            messages.error(self.request, f'Error al importar: {str(e)}')
            return self.form_invalid(form)

        if errores:
            logger.warning(f"Import completed with {len(errores)} errors")
            messages.warning(
                self.request,
                f'Importación completada con {len(errores)} errores. Verifique los logs para más detalles.'
            )

        return super().form_valid(form)


class DescargarPlantillaEnfermerasView(LoginRequiredMixin, View):
    """Descarga plantilla Excel para importar enfermeras"""

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Enfermeras"

        # Headers con estilo
        headers = ['Nombre', 'Email', 'Teléfono', 'DNI', 'Activa']
        ws.append(headers)

        # Aplicar estilo a headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Ejemplo
        ws.append(['María García', 'maria@hospital.com', '600123456', '12345678A', 'Sí'])
        ws.append(['Juan López', 'juan@hospital.com', '600654321', '87654321B', 'Sí'])

        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_enfermeras.xlsx'
        wb.save(response)

        return response


# ========== Tipos de Turno ==========

class TipoTurnoListView(LoginRequiredMixin, ListView):
    """Lista de tipos de turno"""
    model = TipoTurno
    template_name = 'turnos/tipo_turno_list.html'
    context_object_name = 'tipos_turno'

    def get_queryset(self):
        from django.db.models import Count
        return TipoTurno.objects.all().annotate(
            num_configs_count=Count('configuracionplanificacion', distinct=True)
        ).order_by('nombre')


class TipoTurnoCreateView(LoginRequiredMixin, FormMessageMixin, CreateView):
    """Crear tipo de turno"""
    model = TipoTurno
    form_class = TipoTurnoForm
    template_name = 'turnos/tipo_turno_form.html'
    success_url = reverse_lazy('turnos:tipo_turno_lista')
    success_message = 'Tipo de turno creado con éxito.'


class TipoTurnoUpdateView(LoginRequiredMixin, FormMessageMixin, UpdateView):
    """Editar tipo de turno"""
    model = TipoTurno
    form_class = TipoTurnoForm
    template_name = 'turnos/tipo_turno_form.html'
    success_url = reverse_lazy('turnos:tipo_turno_lista')
    success_message = 'Tipo de turno actualizado con éxito.'


class TipoTurnoDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar tipo de turno"""
    model = TipoTurno
    template_name = 'turnos/tipo_turno_confirm_delete.html'
    success_url = reverse_lazy('turnos:tipo_turno_lista')
    context_object_name = 'tipo_turno'

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Verificar si el tipo de turno está en uso en alguna configuración
        num_configs = self.object.configuracionplanificacion.count()
        if num_configs > 0:
            messages.error(
                request,
                f'No se puede eliminar el tipo de turno porque está siendo usado '
                f'en {num_configs} configuraci\u00f3n{"es" if num_configs > 1 else ""}. '
                f'Elimina o modifica esas configuraciones primero.'
            )
            return redirect('turnos:tipo_turno_lista')
        
        messages.success(request, 'Tipo de turno eliminado con éxito.')
        return super().delete(request, *args, **kwargs)


class CrearTurnosPredeterminadosView(LoginRequiredMixin, View):
    """Crea los turnos predeterminados (Mañana, Tarde, Noche)"""

    def post(self, request):
        turnos_default = [
            {
                'nombre': 'MANANA',
                'hora_inicio': time(7, 0),
                'hora_fin': time(15, 0),
                'descripcion': 'Turno de mañana'
            },
            {
                'nombre': 'TARDE',
                'hora_inicio': time(15, 0),
                'hora_fin': time(23, 0),
                'descripcion': 'Turno de tarde'
            },
            {
                'nombre': 'NOCHE',
                'hora_inicio': time(23, 0),
                'hora_fin': time(7, 0),
                'descripcion': 'Turno de noche'
            }
        ]

        creados = 0
        for turno_data in turnos_default:
            if not TipoTurno.objects.filter(nombre=turno_data['nombre']).exists():
                TipoTurno.objects.create(**turno_data)
                creados += 1

        if creados > 0:
            messages.success(request, f'{creados} tipos de turno creados con éxito.')
        else:
            messages.info(request, 'Los turnos predeterminados ya existen.')

        return redirect('turnos:tipo_turno_lista')


# ========== Planillas ==========

class PlanillaListView(LoginRequiredMixin, SearchMixin, PaginationMixin, ListView):
    """Lista de planillas"""
    model = Planilla
    template_name = 'turnos/planilla_list.html'
    context_object_name = 'planillas'
    paginate_by = 12

    search_fields = ['nombre', 'descripcion']

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.select_related('ejecucion__configuracion').order_by('-fecha_inicio')


class PlanillaDetailView(DetailView):
    model = Planilla
    template_name = 'turnos/planilla_detalle.html'  # o el nombre que uses
    context_object_name = 'planilla'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        planilla = self.object

        # Obtener configuración
        config = planilla.configuracion
        fecha_inicio = config.fecha_inicio
        num_dias = config.num_dias

        # Obtener TODAS las asignaciones
        asignaciones = planilla.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera__nombre')

        # Crear estructura de días
        dias = []
        for i in range(num_dias):
            fecha = fecha_inicio + timedelta(days=i)
            dias.append({
                'fecha': fecha,
                'dia_semana': ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo'][fecha.weekday()]
            })

        # Crear estructura de enfermeras con sus turnos
        # Agrupar asignaciones por enfermera
        asignaciones_por_enfermera = defaultdict(dict)

        for asig in asignaciones:
            fecha_str = asig.fecha.strftime('%Y-%m-%d')
            asignaciones_por_enfermera[asig.enfermera][fecha_str] = asig

        # Construir matriz enfermera x día
        enfermeras_turnos = []
        enfermeras = config.enfermeras.all().order_by('nombre')

        for enfermera in enfermeras:
            turnos_enfermera = []

            for dia in dias:
                fecha_str = dia['fecha'].strftime('%Y-%m-%d')
                asig = asignaciones_por_enfermera[enfermera].get(fecha_str)

                if asig:
                    # Tiene turno asignado
                    turno_color = {
                        'MANANA': 'warning',
                        'TARDE': 'info',
                        'NOCHE': 'dark'
                    }.get(asig.turno.nombre, 'secondary')

                    turnos_enfermera.append({
                        'turno': asig.turno,
                        'turno_color': turno_color,
                        'es_libre': False
                    })
                else:
                    # Día libre
                    turnos_enfermera.append({
                        'turno': None,
                        'turno_color': 'secondary',
                        'es_libre': True
                    })

            enfermeras_turnos.append({
                'enfermera': enfermera,
                'turnos': turnos_enfermera
            })

        # Agregar al contexto
        context['dias'] = dias
        context['enfermeras_turnos'] = enfermeras_turnos
        context['asignaciones'] = asignaciones  # Para otras vistas/tabs

        return context


class PlanillaDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar planilla"""
    model = Planilla
    template_name = 'turnos/planilla_confirm_delete.html'
    success_url = reverse_lazy('turnos:planilla_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Planilla eliminada con éxito.')
        return super().delete(request, *args, **kwargs)


# ========== Reportes ==========

class ReportesView(LoginRequiredMixin, TemplateView):
    """Vista principal de reportes"""
    template_name = 'turnos/reportes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas para los reportes
        context['stats'] = {
            'ultima_actualizacion_carga': timezone.now(),
            'ultima_actualizacion_conflictos': timezone.now(),
            'ultima_actualizacion_tendencias': timezone.now(),
            'conflictos_activos': 0,
            'total_configuraciones': ConfiguracionPlanificacion.objects.count(),
            'total_ejecuciones': Ejecucion.objects.count(),
            'total_enfermeras': Enfermera.objects.count(),
            'total_turnos_asignados': AsignacionTurno.objects.count()
        }

        return context


class ReporteCargaView(LoginRequiredMixin, TemplateView):
    """Reporte de carga de trabajo"""
    template_name = 'turnos/reporte_carga.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Filtros
        fecha_desde = self.request.GET.get('fecha_desde')
        fecha_hasta = self.request.GET.get('fecha_hasta')

        # Estadísticas por enfermera
        enfermeras = Enfermera.objects.filter(activa=True)

        datos_enfermeras = []
        for enfermera in enfermeras:
            asignaciones = AsignacionTurno.objects.filter(enfermera=enfermera)

            if fecha_desde:
                asignaciones = asignaciones.filter(fecha__gte=fecha_desde)
            if fecha_hasta:
                asignaciones = asignaciones.filter(fecha__lte=fecha_hasta)

            turnos_manana = asignaciones.filter(turno__nombre='MANANA').count()
            turnos_tarde = asignaciones.filter(turno__nombre='TARDE').count()
            turnos_noche = asignaciones.filter(turno__nombre='NOCHE').count()
            total_turnos = turnos_manana + turnos_tarde + turnos_noche

            datos_enfermeras.append({
                'enfermera': enfermera,
                'turnos_manana': turnos_manana,
                'turnos_tarde': turnos_tarde,
                'turnos_noche': turnos_noche,
                'total_turnos': total_turnos,
                'horas_totales': total_turnos * 8,  # Aproximación
                'dias_libres': asignaciones.filter(es_dia_libre=True).count(),
                'porcentaje_carga': (total_turnos / 30 * 100) if total_turnos else 0
            })

        context['datos_enfermeras'] = datos_enfermeras
        context['resumen'] = {
            'total_enfermeras': len(datos_enfermeras),
            'total_turnos': sum(d['total_turnos'] for d in datos_enfermeras),
            'horas_totales': sum(d['horas_totales'] for d in datos_enfermeras),
            'promedio_por_enfermera': sum(d['total_turnos'] for d in datos_enfermeras) / len(
                datos_enfermeras) if datos_enfermeras else 0
        }

        # Datos para gráfico
        context['distribucion_turnos'] = [
            sum(d['turnos_manana'] for d in datos_enfermeras),
            sum(d['turnos_tarde'] for d in datos_enfermeras),
            sum(d['turnos_noche'] for d in datos_enfermeras)
        ]

        return context


class ReporteConflictosView(LoginRequiredMixin, TemplateView):
    """Reporte de conflictos en planificaciones"""
    template_name = 'turnos/reporte_conflictos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener última ejecución completada del workspace del usuario
        workspace_id = self.request.session.get('workspace_id')
        if workspace_id:
            try:
                workspace = Workspace.objects.get(id=workspace_id, usuarios=self.request.user)
            except Workspace.DoesNotExist:
                workspace = self.request.user.workspaces.first()
        else:
            workspace = self.request.user.workspaces.first()

        conflictos = []
        resumen = {'total_conflictos': 0, 'severidad_alta': 0, 'severidad_media': 0, 'severidad_baja': 0}

        if workspace:
            ultima_ejecucion = Ejecucion.objects.filter(
                workspace=workspace,
                estado='COMPLETADA'
            ).order_by('-fecha_inicio').first()

            if ultima_ejecucion and ultima_ejecucion.mensajes:
                mensajes = ultima_ejecucion.mensajes
                violaciones = mensajes.get('violaciones', [])
                warnings = mensajes.get('warnings', [])

                for v in violaciones:
                    conflictos.append({
                        'tipo': v.get('tipo', 'restriccion'),
                        'mensaje': v.get('mensaje', str(v)),
                        'severidad': 'alta',
                        'ejecucion_id': ultima_ejecucion.id,
                    })
                    resumen['severidad_alta'] += 1

                for w in warnings:
                    severidad = 'media' if 'cobertura' in str(w).lower() else 'baja'
                    conflictos.append({
                        'tipo': 'warning',
                        'mensaje': str(w),
                        'severidad': severidad,
                        'ejecucion_id': ultima_ejecucion.id,
                    })
                    resumen[f'severidad_{severidad}'] += 1

                resumen['total_conflictos'] = len(conflictos)

        context['conflictos'] = conflictos
        context['resumen'] = resumen

        return context


class ReporteTendenciasView(LoginRequiredMixin, TemplateView):
    """Reporte de tendencias temporales"""
    template_name = 'turnos/reporte_tendencias.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Datos mensuales
        hoy = timezone.now()
        datos_mensuales = []

        for i in range(6):
            mes_inicio = hoy - timedelta(days=30 * i)
            mes_fin = mes_inicio + timedelta(days=30)

            ejecuciones_mes = Ejecucion.objects.filter(
                fecha_inicio__gte=mes_inicio,
                fecha_inicio__lt=mes_fin
            )

            # Calcular duración promedio
            avg_duration_seconds = ejecuciones_mes.annotate(
                duracion_calc=ExpressionWrapper(F('fecha_fin') - F('fecha_inicio'), output_field=DurationField())
            ).aggregate(Avg('duracion_calc'))['duracion_calc__avg']

            avg_duration = avg_duration_seconds.total_seconds() if avg_duration_seconds else 0

            datos_mensuales.append({
                'nombre': mes_inicio.strftime('%B'),
                'total': ejecuciones_mes.count(),
                'exitosas': ejecuciones_mes.filter(estado='COMPLETADA').count(),
                'fallidas': ejecuciones_mes.filter(estado='ERROR').count(),
                'tiempo_promedio': avg_duration,
                'tasa_exito': 92.5
            })

        context['datos_mensuales'] = datos_mensuales

        # KPIs
        context['kpis'] = {
            'crecimiento_ejecuciones': 15.5,
            'tiempo_promedio': 45.3,
            'tasa_exito': 92.5,
            'soluciones_optimas': 78.2
        }

        return context


# ========== Vistas de Usuario ==========

class PerfilView(LoginRequiredMixin, TemplateView):
    """Vista de perfil de usuario"""
    template_name = 'turnos/perfil.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas del usuario
        context['stats'] = {
            'configuraciones': ConfiguracionPlanificacion.objects.filter(creado_por=self.request.user).count(),
            'ejecuciones': Ejecucion.objects.filter(configuracion__creado_por=self.request.user).count(),
            'enfermeras': Enfermera.objects.count()
        }

        return context


class PreferenciasView(LoginRequiredMixin, TemplateView):
    """Vista de preferencias del sistema"""
    template_name = 'turnos/preferencias.html'


class GuardarPreferenciasView(LoginRequiredMixin, View):
    """Guarda las preferencias del usuario"""

    def post(self, request):
        # Guardar preferencias en la sesión o en un modelo de perfil
        request.session['preferencias'] = {
            'notificaciones_email': request.POST.get('notificaciones_email') == 'on',
            'notificaciones_browser': request.POST.get('notificaciones_browser') == 'on',
            'idioma': request.POST.get('idioma', 'es'),
            'tema': request.POST.get('tema', 'light'),
            'trabajadores_default': int(request.POST.get('trabajadores_default', 4)),
            'tiempo_maximo_default': int(request.POST.get('tiempo_maximo_default', 60))
        }

        messages.success(request, 'Preferencias guardadas con éxito.')
        return redirect('turnos:preferencias')


# ========== Vistas de Resultado ==========

class ResultadoCalendarioView(LoginRequiredMixin, DetailView):
    """Vista de resultado en formato calendario"""
    model = Ejecucion
    template_name = 'turnos/resultado_calendario.html'
    context_object_name = 'ejecucion'


class ResultadoEstadisticasView(LoginRequiredMixin, DetailView):
    """Vista de estadísticas del resultado"""
    model = Ejecucion
    template_name = 'turnos/resultado_estadisticas.html'
    context_object_name = 'ejecucion'


class ResultadoTablaView(LoginRequiredMixin, DetailView):
    """Vista de resultado en formato tabla"""
    model = Ejecucion
    template_name = 'turnos/resultado_tabla.html'
    context_object_name = 'ejecucion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ejecucion = self.object

        try:
            planilla = ejecucion.planilla_generada
        except Exception:
            planilla = None

        if not planilla:
            return context

        config = ejecucion.configuracion
        fecha_inicio = config.fecha_inicio
        num_dias = config.num_dias

        DIAS_CORTOS = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']

        # Construir lista de días
        dias = []
        for i in range(num_dias):
            fecha = fecha_inicio + timedelta(days=i)
            dias.append({
                'nombre': DIAS_CORTOS[fecha.weekday()],
                'fecha': fecha,
                'fin_semana': fecha.weekday() >= 5,
            })

        # Agrupar días en semanas (corte Lunes-Domingo)
        semanas = []
        semana_actual = {'dias': 0}
        for dia in dias:
            if dia['nombre'] == 'Lun' and semana_actual['dias'] > 0:
                semanas.append(semana_actual)
                semana_actual = {'dias': 0}
            semana_actual['dias'] += 1
        if semana_actual['dias'] > 0:
            semanas.append(semana_actual)

        # Obtener todas las asignaciones
        asignaciones = planilla.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera__nombre')

        # Agrupar por enfermera
        asignaciones_por_enfermera = defaultdict(dict)
        for asig in asignaciones:
            asignaciones_por_enfermera[asig.enfermera_id][asig.fecha] = asig

        # Construir planilla_por_enfermera
        CLASE_CSS = {
            'MANANA': 'turno-manana',
            'TARDE': 'turno-tarde',
            'NOCHE': 'turno-noche',
        }

        enfermeras = config.enfermeras.all().order_by('nombre')
        planilla_por_enfermera = []
        totales_por_dia = [0] * num_dias
        total_general = 0
        dias_libres_totales = 0
        horas_totales = 0

        for enfermera in enfermeras:
            asignaciones_enf = asignaciones_por_enfermera.get(enfermera.id, {})
            fila_asignaciones = []
            total_turnos_enf = 0

            for idx, dia in enumerate(dias):
                fecha = dia['fecha']
                asig = asignaciones_enf.get(fecha)

                if asig and asig.turno and not asig.es_dia_libre and not asig.turno.es_sustituto_libre:
                    turno = asig.turno
                    turno_display = turno.codigo_display() if hasattr(turno, 'codigo_display') else turno.nombre
                    horario = f"{turno.hora_inicio.strftime('%H:%M')}-{turno.hora_fin.strftime('%H:%M')}"
                    clase_css = CLASE_CSS.get(turno.nombre, '')
                    detalle = f"{turno.nombre} ({horario})"
                    total_turnos_enf += 1
                    totales_por_dia[idx] += 1
                    total_general += 1
                    horas_totales += turno.duracion_horas
                    fila_asignaciones.append({
                        'turno': turno,
                        'turno_display': turno_display,
                        'horario': horario,
                        'clase_css': clase_css,
                        'detalle': detalle,
                        'fin_semana': dia['fin_semana'],
                    })
                elif asig and asig.turno and asig.turno.es_sustituto_libre:
                    turno = asig.turno
                    dias_libres_totales += 1
                    fila_asignaciones.append({
                        'turno': None,
                        'turno_display': turno.nombre,
                        'horario': '',
                        'clase_css': 'dia-libre',
                        'detalle': f"Libre ({turno.nombre})",
                        'fin_semana': dia['fin_semana'],
                    })
                else:
                    dias_libres_totales += 1
                    fila_asignaciones.append({
                        'turno': None,
                        'turno_display': '',
                        'horario': '',
                        'clase_css': 'dia-libre',
                        'detalle': 'Día libre',
                        'fin_semana': dia['fin_semana'],
                    })

            planilla_por_enfermera.append({
                'enfermera': enfermera,
                'asignaciones': fila_asignaciones,
                'total_turnos': total_turnos_enf,
            })

        num_enfermeras = enfermeras.count() or 1
        context.update({
            'planilla_por_enfermera': planilla_por_enfermera,
            'dias': dias,
            'semanas': semanas,
            'totales_por_dia': totales_por_dia,
            'total_general': total_general,
            'resumen': {
                'total_turnos': total_general,
                'promedio_por_enfermera': round(total_general / num_enfermeras, 1),
                'dias_libres_totales': dias_libres_totales,
                'horas_totales': round(horas_totales, 1),
            },
        })

        return context


class ResultadoCompararView(LoginRequiredMixin, TemplateView):
    """Vista para comparar dos resultados"""
    template_name = 'turnos/resultado_comparar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        ejecucion_a_id = self.request.GET.get('ejecucion_a')
        ejecucion_b_id = self.request.GET.get('ejecucion_b')

        if ejecucion_a_id and ejecucion_b_id:
            context['ejecucion_a'] = get_object_or_404(Ejecucion, pk=ejecucion_a_id)
            context['ejecucion_b'] = get_object_or_404(Ejecucion, pk=ejecucion_b_id)

        context['ejecuciones'] = Ejecucion.objects.filter(estado='COMPLETADA').order_by('-fecha_inicio')[:20]

        return context


# ========== Vistas de Utilidad ==========

class MaintenanceView(TemplateView):
    """Vista de mantenimiento"""
    template_name = 'turnos/maintenance.html'


# ========== Exportaciones ==========

class ExportarEjecucionExcelView(LoginRequiredMixin, View):
    """Exporta una ejecución a Excel (7 hojas con planilla horizontal)"""

    def get(self, request, pk):
        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        try:
            _ = ejecucion.planilla_generada
        except Exception:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        try:
            buffer = generar_excel_planilla(ejecucion)
            response = FileResponse(
                buffer,
                as_attachment=True,
                filename=f'planilla_{ejecucion.id}.xlsx',
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            return response
        except Exception as e:
            logger.error(f"Error generating Excel: {str(e)}", exc_info=True)
            messages.error(request, f'Error al generar Excel: {str(e)}')
            return redirect('turnos:ejecucion_detalle', pk=pk)


class ExportarEjecucionPDFView(LoginRequiredMixin, View):
    """Exporta una ejecución a PDF con formato de planilla horizontal (matriz enfermeras x días)"""

    DIAS_LETRAS = ['L', 'M', 'X', 'J', 'V', 'S', 'D']

    def get(self, request, pk):
        from django.template.loader import render_to_string
        from weasyprint import HTML

        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        try:
            planilla = ejecucion.planilla_generada
        except Exception:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        context = self._build_matrix_context(ejecucion, planilla)

        html_string = render_to_string('turnos/pdf/planilla.html', context)
        html = HTML(string=html_string)
        pdf = html.write_pdf()

        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.pdf'
        return response

    def _build_matrix_context(self, ejecucion, planilla):
        config = ejecucion.configuracion
        fecha_inicio = config.fecha_inicio
        num_dias = config.num_dias

        # Construir encabezados de días
        dias = []
        for i in range(num_dias):
            fecha = fecha_inicio + timedelta(days=i)
            dias.append({
                'numero': fecha.day,
                'letra': self.DIAS_LETRAS[fecha.weekday()],
                'fecha': fecha,
                'fin_semana': fecha.weekday() >= 5,
            })

        # Obtener asignaciones agrupadas por enfermera
        asignaciones = planilla.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('enfermera__nombre', 'fecha')

        asignaciones_por_enfermera = defaultdict(dict)
        for asig in asignaciones:
            asignaciones_por_enfermera[asig.enfermera_id][asig.fecha] = asig

        # Mapeo tipo_celda -> código corto
        CODIGOS_CELDA = {
            'VACACIONES': 'V',
            'PERMISO': 'P',
            'BAJA': 'B',
            'FORMACION': 'F',
            'ASIGNACION_FIJA': 'AF',
        }
        CLASES_CELDA = {
            'VACACIONES': 'vacaciones',
            'PERMISO': 'permiso',
            'BAJA': 'baja',
            'FORMACION': 'formacion',
            'ASIGNACION_FIJA': 'fija',
        }

        # Construir matriz
        enfermeras = config.enfermeras.all().order_by('nombre')
        matriz = []
        totales_por_dia = [0] * num_dias
        total_general = 0
        tiene_vacaciones = False
        tiene_bajas = False
        tiene_formacion = False

        for enfermera in enfermeras:
            asignaciones_enf = asignaciones_por_enfermera.get(enfermera.id, {})
            celdas = []
            total_enf = 0

            for idx, dia in enumerate(dias):
                fecha = dia['fecha']
                asig = asignaciones_enf.get(fecha)

                if asig and not asig.es_dia_libre and asig.turno:
                    codigo = asig.turno.codigo_display()
                    clase = asig.turno.nombre.lower().replace('Ñ', 'n').replace('mañana', 'm').replace('tarde', 't').replace('noche', 'n')
                    # Usar códigos CSS simples
                    clase_map = {'MANANA': 'm', 'TARDE': 't', 'NOCHE': 'n'}
                    clase = clase_map.get(asig.turno.nombre, 'm')
                    total_enf += 1
                    totales_por_dia[idx] += 1
                    total_general += 1
                elif asig and asig.tipo_celda in CODIGOS_CELDA:
                    codigo = CODIGOS_CELDA[asig.tipo_celda]
                    clase = CLASES_CELDA[asig.tipo_celda]
                    if asig.tipo_celda == 'VACACIONES':
                        tiene_vacaciones = True
                    elif asig.tipo_celda == 'BAJA':
                        tiene_bajas = True
                    elif asig.tipo_celda == 'FORMACION':
                        tiene_formacion = True
                else:
                    codigo = '—'
                    clase = 'libre'

                celdas.append({
                    'codigo': codigo,
                    'clase': clase,
                    'fin_semana': dia['fin_semana'],
                })

            matriz.append({
                'enfermera': enfermera.nombre,
                'celdas': celdas,
                'total': total_enf,
            })

        # Formato de período
        fecha_fin = fecha_inicio + timedelta(days=num_dias - 1)
        meses_es = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]
        periodo = f"{meses_es[fecha_inicio.month - 1]} - {fecha_inicio.year}"

        return {
            'ejecucion': ejecucion,
            'planilla': planilla,
            'periodo': periodo,
            'nombre_configuracion': config.nombre,
            'fecha_inicio_str': fecha_inicio.strftime('%d/%m/%Y'),
            'fecha_fin_str': fecha_fin.strftime('%d/%m/%Y'),
            'fecha_generacion': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'dias': dias,
            'matriz': matriz,
            'totales_por_dia': totales_por_dia,
            'total_general': total_general,
            'tiene_vacaciones': tiene_vacaciones,
            'tiene_bajas': tiene_bajas,
            'tiene_formacion': tiene_formacion,
        }


class ExportarEjecucionCSVView(LoginRequiredMixin, View):
    """Exporta una ejecución a CSV"""

    def get(self, request, pk):
        import csv

        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla_generada:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.csv'

        writer = csv.writer(response)
        writer.writerow(['Enfermera', 'Fecha', 'Turno', 'Horario', 'Es Día Libre'])

        asignaciones = ejecucion.planilla_generada.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera')

        for asignacion in asignaciones:
            if asignacion.es_dia_libre:
                turno_info = 'Libre'
                horario = '-'
            else:
                turno_info = asignacion.turno.nombre
                horario = f"{asignacion.turno.hora_inicio.strftime('%H:%M')} - {asignacion.turno.hora_fin.strftime('%H:%M')}"

            writer.writerow([
                asignacion.enfermera.nombre,
                asignacion.fecha.strftime('%d/%m/%Y'),
                turno_info,
                horario,
                'Sí' if asignacion.es_dia_libre else 'No'
            ])

        return response


class ExportarEjecucionJSONView(LoginRequiredMixin, View):
    """Exporta una ejecución a JSON"""

    def get(self, request, pk):
        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla_generada:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        data = {
            'ejecucion_id': ejecucion.id,
            'configuracion': ejecucion.configuracion.nombre,
            'fecha_inicio': ejecucion.fecha_inicio.isoformat(),
            'estado': ejecucion.estado,
            'es_optima': ejecucion.es_optima,
            'penalizacion_total': ejecucion.penalizacion_total,
            'planilla': {
                'nombre': ejecucion.planilla_generada.nombre,
                'fecha_inicio': ejecucion.planilla_generada.fecha_inicio.isoformat(),
                'fecha_fin': ejecucion.planilla_generada.fecha_fin.isoformat(),
                'asignaciones': []
            }
        }

        asignaciones = ejecucion.planilla_generada.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera')

        for asignacion in asignaciones:
            data['planilla']['asignaciones'].append({
                'enfermera': asignacion.enfermera.nombre,
                'fecha': asignacion.fecha.isoformat(),
                'turno': asignacion.turno.nombre if asignacion.turno else None,
                'es_dia_libre': asignacion.es_dia_libre
            })

        response = JsonResponse(data, json_dumps_params={'indent': 2})
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.json'

        return response


class ExportarEjecucionICalView(LoginRequiredMixin, View):
    """Exporta una ejecución a formato iCalendar"""

    def get(self, request, pk):
        from icalendar import Calendar, Event


        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla_generada:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        cal = Calendar()
        cal.add('prodid', '-//Sistema de Planificación de Turnos//ES')
        cal.add('version', '2.0')
        cal.add('X-WR-CALNAME', ejecucion.configuracion.nombre)

        asignaciones = ejecucion.planilla_generada.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha')

        for asignacion in asignaciones:
            if not asignacion.es_dia_libre:
                event = Event()
                event.add('summary', f"{asignacion.enfermera.nombre} - {asignacion.turno.nombre}")

                # Calcular fecha/hora inicio y fin
                dt_inicio = datetime.combine(asignacion.fecha, asignacion.turno.hora_inicio)
                dt_fin = datetime.combine(asignacion.fecha, asignacion.turno.hora_fin)

                # Si el turno cruza medianoche
                if dt_fin <= dt_inicio:
                    dt_fin += timedelta(days=1)

                event.add('dtstart', dt_inicio)
                event.add('dtend', dt_fin)
                event.add('description', f"Turno asignado: {asignacion.turno.nombre}")

                cal.add_component(event)

        response = HttpResponse(cal.to_ical(), content_type='text/calendar')
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.ics'

        return response


class ExportarPlanillaExcelView(LoginRequiredMixin, View):
    """Exporta una planilla específica a Excel"""

    def get(self, request, pk):
        planilla = get_object_or_404(Planilla, pk=pk)
        # Reutilizar la lógica de exportación de ejecución
        # pero usando la planilla directamente
        return ExportarEjecucionExcelView().get(request, planilla.ejecucion.id)


class ExportarPlanillaPDFView(LoginRequiredMixin, View):
    """Exporta una planilla específica a PDF"""

    def get(self, request, pk):
        planilla = get_object_or_404(Planilla, pk=pk)
        return ExportarEjecucionPDFView().get(request, planilla.ejecucion.id)

class ExportarConfiguracionJSONView(LoginRequiredMixin, DetailView):
    """Vista para exportar configuración en formato JSON"""
    model = ConfiguracionPlanificacion

    def get(self, request, *args, **kwargs):
        configuracion = self.get_object()

        data = {
            'nombre': configuracion.nombre,
            'descripcion': configuracion.descripcion,
            'num_dias': configuracion.num_dias,
            'fecha_inicio': configuracion.fecha_inicio.isoformat() if configuracion.fecha_inicio else None,
            'enfermeras': list(configuracion.enfermeras.values_list('id', flat=True)),
            'turnos': list(configuracion.turnos.values_list('id', flat=True)),
            'demanda_por_turno': configuracion.demanda_por_turno,
            'restricciones_duras': configuracion.restricciones_duras,
            'restricciones_blandas': configuracion.restricciones_blandas,
            'num_trabajadores': configuracion.num_trabajadores,
            'tiempo_maximo_segundos': configuracion.tiempo_maximo_segundos,
            'seed': configuracion.seed,
        }

        response = JsonResponse(data, json_dumps_params={'indent': 2, 'ensure_ascii': False})
        response['Content-Disposition'] = f'attachment; filename="config_{configuracion.pk}.json"'
        return response

class WorkspaceMixin(LoginRequiredMixin):
    def get_current_workspace(self):
        workspace_id = self.request.session.get('workspace_id')
        if workspace_id:
            return get_object_or_404(Workspace, id=workspace_id, usuarios=self.request.user)
        return self.request.user.workspaces.first()

    def get_queryset(self):
        qs = super().get_queryset()
        ws = self.get_current_workspace()
        if ws:
            return qs.filter(workspace=ws)
        return qs.none()


class CambiarWorkspaceView(LoginRequiredMixin, View):
    def post(self, request):
        workspace_id = request.POST.get('workspace_id')
        ws = get_object_or_404(Workspace, id=workspace_id, usuarios=request.user)
        request.session['workspace_id'] = ws.id
        return JsonResponse({'success': True, 'workspace': ws.nombre})


# ══════════════════════════════════════════════════════════════
# VISTAS PARA GESTIONAR RESTRICCIONES SACYL
# ══════════════════════════════════════════════════════════════

class ConfiguracionRestriccionesView(LoginRequiredMixin, TemplateView):
    """Vista para editar restricciones de una configuración"""
    template_name = 'turnos/restricciones_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        config_id = self.kwargs.get('pk')

        try:
            config = ConfiguracionPlanificacion.objects.get(pk=config_id)
            context['config'] = config
            context['restricciones_duras'] = config.restricciones_duras or []
            context['restricciones_blandas'] = config.restricciones_blandas or []

            # Importar formularios si existen
            try:
                from .forms import (
                    RestriccionDuraForm,
                    RestriccionBlandaForm,
                    CargarRestriccionesSACYLForm
                )
                context['form_dura'] = RestriccionDuraForm()
                context['form_blanda'] = RestriccionBlandaForm()
                context['form_cargar'] = CargarRestriccionesSACYLForm()
            except ImportError:
                logger.warning("Formularios de restricciones no disponibles")

        except ConfiguracionPlanificacion.DoesNotExist:
            messages.error(self.request, 'Configuración no encontrada')

        return context

    def post(self, request, *args, **kwargs):
        config_id = self.kwargs.get('pk')
        action = request.POST.get('action')
        logger.info(f"POST en ConfiguracionRestriccionesView. Config ID: {config_id}, Acción: {action}")

        try:
            config = ConfiguracionPlanificacion.objects.get(pk=config_id)
        except ConfiguracionPlanificacion.DoesNotExist:
            logger.error(f"Configuración con ID {config_id} no encontrada.")
            messages.error(request, 'Configuración no encontrada')
            return redirect('turnos:config_lista')

        try:
            # Añadir restricción dura
            if action == 'add_dura':
                logger.debug("Acción: add_dura")
                try:
                    from .forms import RestriccionDuraForm
                    form = RestriccionDuraForm(request.POST)
                    if form.is_valid():
                        rd = config.restricciones_duras or []
                        if not isinstance(rd, list): # Defensive check
                            logger.warning(f"restricciones_duras no era una lista, se reinicia. Contenido: {config.restricciones_duras}")
                            rd = []
                        nueva_restriccion = form.to_dict()
                        rd.append(nueva_restriccion)
                        config.restricciones_duras = rd
                        config.save()
                        logger.info(f"Restricción dura añadida a config {config_id}: {nueva_restriccion}")
                        messages.success(request, f'Restricción dura "{form.cleaned_data["nombre"]}" añadida')
                    else:
                        logger.warning(f"Formulario de restricción dura inválido para config {config_id}: {form.errors.as_json()}")
                        for field, errors in form.errors.items():
                            for error in errors:
                                messages.error(request, f'{field}: {error}')
                except ImportError:
                    logger.error("No se pudo importar RestriccionDuraForm.")
                    messages.error(request, 'Formulario de restricciones no disponible')

            # Añadir restricción blanda
            elif action == 'add_blanda':
                logger.debug("Acción: add_blanda")
                try:
                    from .forms import RestriccionBlandaForm
                    form = RestriccionBlandaForm(request.POST)
                    if form.is_valid():
                        rb = config.restricciones_blandas or []
                        if not isinstance(rb, list): # Defensive check
                            logger.warning(f"restricciones_blandas no era una lista, se reinicia. Contenido: {config.restricciones_blandas}")
                            rb = []
                        nueva_restriccion = form.to_dict()
                        rb.append(nueva_restriccion)
                        config.restricciones_blandas = rb
                        config.save()
                        logger.info(f"Restricción blanda añadida a config {config_id}: {nueva_restriccion}")
                        messages.success(request, f'Restricción blanda "{form.cleaned_data["nombre"]}" añadida')
                    else:
                        logger.warning(f"Formulario de restricción blanda inválido para config {config_id}: {form.errors.as_json()}")
                        for field, errors in form.errors.items():
                            for error in errors:
                                messages.error(request, f'{field}: {error}')
                except ImportError:
                    logger.error("No se pudo importar RestriccionBlandaForm.")
                    messages.error(request, 'Formulario de restricciones no disponible')

            # Cargar preset SACYL
            elif action == 'cargar_sacyl':
                logger.debug("Acción: cargar_sacyl")
                try:
                    from .forms import CargarRestriccionesSACYLForm
                    form = CargarRestriccionesSACYLForm(request.POST, request.FILES)
                    if form.is_valid():
                        preset = form.cleaned_data['preset']
                        logger.info(f"Cargando preset SACYL '{preset}' en config {config_id}")

                        if preset == 'basico':
                            # Cargar preset básico hardcoded
                            config.restricciones_duras = [
                                {"id": "RD006", "nombre": "descanso_minimo_12h", "tipo": "descanso", "obligatorio": True, "parametros": {"minimo_horas": 12}, "descripcion": "Descanso mínimo 12h entre jornadas"},
                                {"id": "RD019", "nombre": "cobertura_minima", "tipo": "cobertura", "obligatorio": True, "parametros": {}, "descripcion": "Cobertura mínima por turno"},
                                {"id": "RD020", "nombre": "no_solapamiento", "tipo": "asignacion", "obligatorio": True, "parametros": {}, "descripcion": "Una enfermera, un turno por día"}
                            ]
                            config.restricciones_blandas = [
                                {"id": "RB001", "nombre": "equidad_turnos", "tipo": "equidad", "peso": 100, "parametros": {}, "descripcion": "Distribución equitativa de turnos"}
                            ]
                            config.save()
                            logger.info(f"Preset básico SACYL guardado en config {config_id}.")
                            messages.success(request, 'Preset básico SACYL cargado')

                        elif preset == 'custom' and 'json_data' in form.cleaned_data:
                            data = form.cleaned_data['json_data']
                            logger.debug(f"Cargando restricciones desde JSON: {json.dumps(data, separators=(',', ':'), ensure_ascii=False)}")
                            if 'restricciones_duras' in data:
                                config.restricciones_duras = data['restricciones_duras']
                                logger.info(f"Restricciones duras actualizadas desde JSON para config {config_id}.")
                            if 'restricciones_blandas' in data:
                                config.restricciones_blandas = data['restricciones_blandas']
                                logger.info(f"Restricciones blandas actualizadas desde JSON para config {config_id}.")
                            config.save()
                            messages.success(request, 'Restricciones personalizadas cargadas desde JSON')
                    else:
                        logger.warning(f"Formulario CargarRestriccionesSACYLForm inválido para config {config_id}: {form.errors.as_json()}")
                        for error in form.non_field_errors():
                            messages.error(request, error)
                except ImportError:
                    logger.error("No se pudo importar CargarRestriccionesSACYLForm.")
                    messages.error(request, 'Formulario de carga SACYL no disponible')

            # Eliminar restricción
            elif action == 'delete_dura':
                idx_str = request.POST.get('index', '-1')
                logger.debug(f"Acción: delete_dura, index: {idx_str}")
                idx = int(idx_str)
                if idx >= 0:
                    rd = config.restricciones_duras or []
                    if isinstance(rd, list) and idx < len(rd):
                        eliminada = rd.pop(idx)
                        config.restricciones_duras = rd
                        config.save()
                        logger.info(f"Restricción dura eliminada de config {config_id}. Restricción: {eliminada}")
                        messages.success(request, 'Restricción dura eliminada')
                    else:
                        logger.warning(f"Índice de restricción dura inválido o fuera de rango para config {config_id}. Index: {idx}, Total: {len(rd)}")

            elif action == 'delete_blanda':
                idx_str = request.POST.get('index', '-1')
                logger.debug(f"Acción: delete_blanda, index: {idx_str}")
                idx = int(idx_str)
                if idx >= 0:
                    rb = config.restricciones_blandas or []
                    if isinstance(rb, list) and idx < len(rb):
                        eliminada = rb.pop(idx)
                        config.restricciones_blandas = rb
                        config.save()
                        logger.info(f"Restricción blanda eliminada de config {config_id}. Restricción: {eliminada}")
                        messages.success(request, 'Restricción blanda eliminada')
                    else:
                        logger.warning(f"Índice de restricción blanda inválido o fuera de rango para config {config_id}. Index: {idx}, Total: {len(rb)}")

            else:
                logger.warning(f"Acción desconocida '{action}' en ConfiguracionRestriccionesView.")

        except Exception as e:
            logger.exception(f"Error inesperado al procesar la acción '{action}' para la config {config_id}")
            messages.error(request, f"Se produjo un error inesperado: {e}")

        return redirect('turnos:config_restricciones', pk=config_id)


# ========================================
# VISTAS DE EXPORTACIÓN
# ========================================
# =========================================================================
# DESCARGAS DE EXPORTACIÓN
# =========================================================================

class DescargarExcelView(LoginRequiredMixin, View):
    """Descarga Excel de ejecución"""

    def get(self, request, pk):
        """Descarga el Excel"""
        try:
            ejecucion = Ejecucion.objects.get(pk=pk)
            logger.info(f"Excel download requested for execution {pk}")

            # Generar Excel
            buffer = generar_excel_planilla(ejecucion)

            # Crear respuesta
            response = FileResponse(
                buffer,
                as_attachment=True,
                filename=f'planificacion_{ejecucion.configuracion.nombre}_{ejecucion.id}.xlsx',
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            logger.info(f"Excel downloaded successfully for execution {pk}")
            return response

        except Ejecucion.DoesNotExist:
            logger.error(f"Execution {pk} not found")
            messages.error(request, "Ejecución no encontrada")
            return redirect('turnos:ejecucion_lista')
        except Exception as e:
            logger.error(f"Error generating Excel: {str(e)}", exc_info=True)
            messages.error(request, f"Error al generar Excel: {str(e)}")
            return redirect('turnos:ejecucion_detalle', pk=pk)


class DescargarPDFView(LoginRequiredMixin, View):
    """Descarga PDF de ejecución"""

    def get(self, request, pk):
        """Descarga el PDF"""
        try:
            ejecucion = Ejecucion.objects.get(pk=pk)
            logger.info(f"PDF download requested for execution {pk}")

            # Generar PDF
            buffer = generar_pdf_planilla(ejecucion)

            # Crear respuesta
            response = FileResponse(
                buffer,
                as_attachment=True,
                filename=f'planificacion_{ejecucion.configuracion.nombre}_{ejecucion.id}.pdf',
                content_type='application/pdf'
            )

            logger.info(f"PDF downloaded successfully for execution {pk}")
            return response

        except Ejecucion.DoesNotExist:
            logger.error(f"Execution {pk} not found")
            messages.error(request, "Ejecución no encontrada")
            return redirect('turnos:ejecucion_lista')
        except Exception as e:
            logger.error(f"Error generating PDF: {str(e)}", exc_info=True)
            messages.error(request, f"Error al generar PDF: {str(e)}")
            return redirect('turnos:ejecucion_detalle', pk=pk)


class DescargarCSVView(LoginRequiredMixin, View):
    """Descarga CSV de ejecución"""

    def get(self, request, pk):
        """Descarga el CSV"""
        try:
            ejecucion = Ejecucion.objects.get(pk=pk)
            logger.info(f"CSV download requested for execution {pk}")

            # Generar CSV
            buffer = generar_csv_planilla(ejecucion)

            # Crear respuesta
            response = FileResponse(
                buffer,
                as_attachment=True,
                filename=f'planificacion_{ejecucion.configuracion.nombre}_{ejecucion.id}.csv',
                content_type='text/csv; charset=utf-8-sig'
            )

            logger.info(f"CSV downloaded successfully for execution {pk}")
            return response

        except Ejecucion.DoesNotExist:
            logger.error(f"Execution {pk} not found")
            messages.error(request, "Ejecución no encontrada")
            return redirect('turnos:ejecucion_lista')
        except Exception as e:
            logger.error(f"Error generating CSV: {str(e)}", exc_info=True)
            messages.error(request, f"Error al generar CSV: {str(e)}")
            return redirect('turnos:ejecucion_detalle', pk=pk)


class DescargarJSONView(LoginRequiredMixin, View):
    """Descarga JSON de ejecución"""

    def get(self, request, pk):
        """Descarga el JSON"""
        try:
            ejecucion = Ejecucion.objects.get(pk=pk)
            logger.info(f"JSON download requested for execution {pk}")

            # Generar JSON
            buffer = generar_json_planilla(ejecucion)

            # Crear respuesta
            response = FileResponse(
                buffer,
                as_attachment=True,
                filename=f'planificacion_{ejecucion.configuracion.nombre}_{ejecucion.id}.json',
                content_type='application/json'
            )

            logger.info(f"JSON downloaded successfully for execution {pk}")
            return response

        except Ejecucion.DoesNotExist:
            logger.error(f"Execution {pk} not found")
            messages.error(request, "Ejecución no encontrada")
            return redirect('turnos:ejecucion_lista')
        except Exception as e:
            logger.error(f"Error generating JSON: {str(e)}", exc_info=True)
            messages.error(request, f"Error al generar JSON: {str(e)}")
            return redirect('turnos:ejecucion_detalle', pk=pk)


class DescargarICalView(LoginRequiredMixin, View):
    """Descarga iCalendar de ejecución"""

    def get(self, request, pk):
        """Descarga el iCal"""
        try:
            ejecucion = Ejecucion.objects.get(pk=pk)
            logger.info(f"iCal download requested for execution {pk}")

            # Generar iCal
            buffer = generar_ical_planilla(ejecucion)

            # Crear respuesta
            response = FileResponse(
                buffer,
                as_attachment=True,
                filename=f'planificacion_{ejecucion.configuracion.nombre}_{ejecucion.id}.ics',
                content_type='text/calendar'
            )

            logger.info(f"iCal downloaded successfully for execution {pk}")
            return response

        except Ejecucion.DoesNotExist:
            logger.error(f"Execution {pk} not found")
            messages.error(request, "Ejecución no encontrada")
            return redirect('turnos:ejecucion_lista')
        except Exception as e:
            logger.error(f"Error generating iCal: {str(e)}", exc_info=True)
            messages.error(request, f"Error al generar iCalendar: {str(e)}")
            return redirect('turnos:ejecucion_detalle', pk=pk)


class DescargarEnfermerasExcelView(LoginRequiredMixin, View):
    """Descarga lista de enfermeras en Excel"""

    def get(self, request):
        """Descarga el Excel de enfermeras"""
        try:
            logger.info(f"Enfermeras Excel download requested by user {request.user.username}")

            # Filtrar por workspace del usuario
            workspace_id = request.session.get('workspace_id')
            if workspace_id:
                workspace = get_object_or_404(Workspace, id=workspace_id, usuarios=request.user)
            else:
                workspace = request.user.workspaces.first()

            if workspace:
                enfermeras = Enfermera.objects.filter(workspace=workspace)
            else:
                enfermeras = Enfermera.objects.none()

            buffer = exportar_enfermeras_excel(enfermeras)

            response = FileResponse(
                buffer,
                as_attachment=True,
                filename=f'enfermeras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            logger.info(f"Enfermeras Excel downloaded successfully")
            return response

        except Exception as e:
            logger.error(f"Error generating enfermeras Excel: {str(e)}", exc_info=True)
            messages.error(request, f"Error al generar Excel: {str(e)}")
            return redirect('turnos:dashboard')
